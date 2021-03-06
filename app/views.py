from django.http import HttpResponseRedirect, HttpResponseNotFound, HttpResponse,Http404
from django.urls import reverse
from django.shortcuts import render
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout
from openpyxl import load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from .forms import SE_Form, LoginForm, CustomUserChangeForm
from .models import RegionModel
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from django.shortcuts import get_object_or_404

regions_names = ['Абзелиловский район РБ', 'Агидель РБ', 'Альшеевский район РБ', 'Архангельский район РБ',
                 'Аскинский район РБ', 'Аургазинский район РБ', 'Баймакский район РБ', 'Бакалинский район РБ',
                 'Балтачевский район РБ', 'Белебеевский район РБ', 'Белокатайский район РБ', 'Белорецкий район РБ',
                 'Бижбулякский район РБ', 'Бирский район РБ', 'Благоварский район РБ', 'Благовещенский район РБ',
                 'Буздякский район РБ', 'Бураевский район РБ', 'Бурзянский район РБ', 'Гафурийский район РБ',
                 'Давлекановский район РБ', 'Дуванский район РБ', 'Дюртюлинский район РБ', 'Ермекеевский район',
                 'Зианчуринский район РБ', 'Зилаирский район РБ', 'Иглинский район РБ', 'Илишевский район РБ',
                 'Ишимбайский район РБ', 'Калтасинский район РБ', 'Караидельский район РБ', 'Кармаскалинский район РБ',
                 'Кигинский район РБ', 'Краснокамский район РБ', 'Кугарчинский район РБ', 'Кумертау',
                 'Кушнаренковский район РБ', 'Куюргазинский район РБ', 'Межгорье', 'Мелеузовский район РБ',
                 'Мечетлинский район РБ', 'Мишкинский район РБ', 'Миякинский район РБ', 'Нефтекамск',
                 'Нуримановский район РБ', 'Октябрьский', 'Салават', 'Салаватский район РБ',
                 'Сибай', 'Стерлибашевский район РБ', 'Стерлитамак', 'Стерлитамакский район РБ',
                 'Татышлинский район РБ', 'Туймазинский район РБ', 'Уфа', 'Уфимский район РБ', 'Учалинский район РБ',
                 'Федоровский район РБ', 'Хайбуллинский район РБ', 'Чекмагушевский район РБ', 'Чишминский район РБ',
                 'Шаранский район РБ', 'Янаульский район РБ']

short_regions_names = ['abzelil', 'agidel', 'alsheev', 'archang', 'askinsk', 'aurgazin', 'baymak', 'bakalin', 'baltach',
                       'belebeev', 'belokatay', 'belorezk', 'bizhbul', 'birsk', 'blagovar', 'blagovesch', 'buzdyak',
                       'buraev', 'burzyan', 'gafur', 'davlekan', 'duvansk', 'dyurtyulin', 'ermekeev', 'zianchurin',
                       'zilairsk', 'iglinsk', 'ilishevsk', 'ishimbaysk', 'kaltasinsk', 'karaidelsks', 'karmaskalin',
                       'kiginsk', 'krasnokamsk', 'kugarchinsk', 'kumertau', 'kushnarenk', 'kuyurgazinsk', 'mezhgorie',
                       'meleuz', 'mechetlin', 'mishkin', 'miyakin', 'neftekamsk', 'nurimanovsk', 'oktabrsks',
                       'slavat', 'salavatskiy', 'sibi', 'sterlibash', 'sterlitamak', 'sterlitamakskiy',
                       'tatyshlin', 'tuymazin', 'ufa', 'ufimsk', 'uchalinsk', 'fedorovsk', 'haybullinsk',
                       'chekmagush', 'chishminsk', 'sharanck', 'yanaulsk']


short_service_names = ['residential_premises', 'housing_transfer', 'advertising_structures', 'capital_construction',
                       'preschool_education', 'school_education', 'needing_premises', 'town_planning',
                       'archive_reference', 'land_schemes', 'land_sale', 'land_lease', 'ownership_right',
                       'municipal_property_lease', 'free_land_provision']

full_service_names = ['Согласование проведения переустройства и (или) перепланировки помещения в многоквартирном доме',
                      'Выдача решения о переводе или об отказе в переводе жилого помещения в нежилое помещение или нежилого помещения в жилое помещение',
                      'Выдача разрешения на установку и эксплуатацию рекламной конструкции ',
                      'Выдача разрешения на строительство объекта капитального строительства',
                      'Постановка на учет и зачисление детей в образовательные учреждения, реализующие образовательную программу дошкольного образования (детские сады)',
                      'Зачисление детей в муниципальные общеобразовательные учреждения',
                      'Принятие на учет граждан в качестве нуждающихся в жилых помещениях',
                      'Выдача градостроительных планов земельных участков',
                      'Предоставление архивных справок, архивных копий, архивных выписок, информационных писем, связанных с реализацией законных прав и свобод граждан и исполнением государственными органами и органами местного самоуправления своих полномочий',
                      'Утверждение схемы расположения земельного участка или земельных участков на кадастровом плане территории',
                      'Продажа земельных участков, находящихся в муниципальной собственности муниципального образования или государственная собственность на которые не разграничена, на которых расположены здания, сооружения, собственникам таких зданий, сооружений либо помещений в них',
                      'Предоставление в аренду земельных участков, находящихся в муниципальной собственности муниципального образования или государственная собственность на которые не разграничена, без проведения торгов',
                      'Выдача копий архивных документов, подтверждающих право на владение землей',
                      'Предоставление муниципального имущества (за исключением земельных участков) в аренду, безвозмездное пользование, доверительное управление без проведения конкурсов или аукционов',
                      'Предоставление земельного участка, находящегося в муниципальной собственности муниципального образования или государственная собственность на который не разграничена, гражданам в собственность бесплатно для индивидуального жилищного строительства']

MONTHS = ['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']

MONTH_NUMBERS = ['01','02','03','04','05','06','07','08','09','10','11','12']

class dotDict(dict):
    """dot.notation access to dictionary attributes"""
    __getattr__ = dict.get
    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__

def get_not_sent(month,year):
    objects = []
    for name in regions_names:
        try:
            get_object_or_404(RegionModel, region_name=name, month=month, year=year)

        except Http404:
            objects.append(name)

    return objects


def get_with_troubles(month, year):
    objects = dict()
    objects.update({'residential_premises': []})
    objects.update({'housing_transfer': []})
    objects.update({'advertising_structures': []})
    objects.update({'capital_construction': []})
    objects.update({'preschool_education': []})
    objects.update({'school_education': []})
    objects.update({'needing_premises': []})
    objects.update({'town_planning': []})
    objects.update({'archive_reference': []})
    objects.update({'land_schemes': []})
    objects.update({'land_sale': []})
    objects.update({'land_lease': []})
    objects.update({'ownership_right': []})
    objects.update({'municipal_property_lease': []})
    objects.update({'free_land_provision': []})
    for name in regions_names:
        try:
            obj = get_object_or_404(RegionModel, region_name=name,month=month,year=year)
            if (obj.residential_premises_has_advanced_appointment_comment != 'Да' and
                obj.residential_premises_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.residential_premises_has_advanced_appointment_comment != '') or \
                (obj.residential_premises_has_btn_get_service_comment != 'Да' and
                 obj.residential_premises_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.residential_premises_has_btn_get_service_comment != '') or \
                    (obj.residential_premises_has_reglament_comment != 'Да' and
                     obj.residential_premises_has_reglament_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_reglament_comment != '') or \
                    (obj.residential_premises_has_estimation_quality_comment != 'Да' and
                     obj.residential_premises_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_estimation_quality_comment != '') or \
                    (obj.residential_premises_connected_to_fgis_do_comment != 'Да' and
                     obj.residential_premises_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.residential_premises_connected_to_fgis_do_comment != '') or \
                    (obj.residential_premises_has_electronic_form_printing_comment != 'Да' and
                     obj.residential_premises_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_electronic_form_printing_comment != '') or \
                    (obj.residential_premises_has_edition_draft_comment != 'Да' and
                     obj.residential_premises_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_edition_draft_comment != '') or \
                    (obj.residential_premises_has_term_of_consideration_comment != 'Да' and
                     obj.residential_premises_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_term_of_consideration_comment != '') or \
                    (obj.residential_premises_has_notif_consider_result_comment != 'Да' and
                     obj.residential_premises_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_notif_consider_result_comment != '') or \
                    (obj.residential_premises_has_causes_of_failure_comment != 'Да' and
                     obj.residential_premises_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_causes_of_failure_comment != '') or \
                    (obj.residential_premises_has_sample_document_comment != 'Да' and
                     obj.residential_premises_has_sample_document_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_sample_document_comment != '') or \
                    (obj.residential_premises_has_document_template_comment != 'Да' and
                     obj.residential_premises_has_document_template_comment != 'Не предусмотрено' and
                     obj.residential_premises_has_document_template_comment != ''):
                objects['residential_premises'].append(obj)

            if (obj.housing_transfer_has_advanced_appointment_comment != 'Да' and
                obj.housing_transfer_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.housing_transfer_has_advanced_appointment_comment != '') or \
                (obj.housing_transfer_has_btn_get_service_comment != 'Да' and
                 obj.housing_transfer_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.housing_transfer_has_btn_get_service_comment != '') or \
                    (obj.housing_transfer_has_reglament_comment != 'Да' and
                     obj.housing_transfer_has_reglament_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_reglament_comment != '') or \
                    (obj.housing_transfer_has_estimation_quality_comment != 'Да' and
                     obj.housing_transfer_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_estimation_quality_comment != '') or \
                    (obj.housing_transfer_connected_to_fgis_do_comment != 'Да' and
                     obj.housing_transfer_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.housing_transfer_connected_to_fgis_do_comment != '') or \
                    (obj.housing_transfer_has_electronic_form_printing_comment != 'Да' and
                     obj.housing_transfer_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_electronic_form_printing_comment != '') or \
                    (obj.housing_transfer_has_edition_draft_comment != 'Да' and
                     obj.housing_transfer_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_edition_draft_comment != '') or \
                    (obj.housing_transfer_has_term_of_consideration_comment != 'Да' and
                     obj.housing_transfer_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_term_of_consideration_comment != '') or \
                    (obj.housing_transfer_has_notif_consider_result_comment != 'Да' and
                     obj.housing_transfer_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_notif_consider_result_comment != '') or \
                    (obj.housing_transfer_has_causes_of_failure_comment != 'Да' and
                     obj.housing_transfer_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_causes_of_failure_comment != '') or \
                    (obj.housing_transfer_has_sample_document_comment != 'Да' and
                     obj.housing_transfer_has_sample_document_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_sample_document_comment != '') or \
                    (obj.housing_transfer_has_document_template_comment != 'Да' and
                     obj.housing_transfer_has_document_template_comment != 'Не предусмотрено' and
                     obj.housing_transfer_has_document_template_comment != ''):
                objects['housing_transfer'].append(obj)

            if (obj.advertising_structures_has_advanced_appointment_comment != 'Да' and
                obj.advertising_structures_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.advertising_structures_has_advanced_appointment_comment != '') or \
                (obj.advertising_structures_has_btn_get_service_comment != 'Да' and
                 obj.advertising_structures_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.advertising_structures_has_btn_get_service_comment != '') or \
                    (obj.advertising_structures_has_reglament_comment != 'Да' and
                     obj.advertising_structures_has_reglament_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_reglament_comment != '') or \
                    (obj.advertising_structures_has_estimation_quality_comment != 'Да' and
                     obj.advertising_structures_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_estimation_quality_comment != '') or \
                    (obj.advertising_structures_connected_to_fgis_do_comment != 'Да' and
                     obj.advertising_structures_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.advertising_structures_connected_to_fgis_do_comment != '') or \
                    (obj.advertising_structures_has_electronic_form_printing_comment != 'Да' and
                     obj.advertising_structures_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_electronic_form_printing_comment != '') or \
                    (obj.advertising_structures_has_edition_draft_comment != 'Да' and
                     obj.advertising_structures_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_edition_draft_comment != '') or \
                    (obj.advertising_structures_has_term_of_consideration_comment != 'Да' and
                     obj.advertising_structures_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_term_of_consideration_comment != '') or \
                    (obj.advertising_structures_has_notif_consider_result_comment != 'Да' and
                     obj.advertising_structures_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_notif_consider_result_comment != '') or \
                    (obj.advertising_structures_has_causes_of_failure_comment != 'Да' and
                     obj.advertising_structures_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_causes_of_failure_comment != '') or \
                    (obj.advertising_structures_has_sample_document_comment != 'Да' and
                     obj.advertising_structures_has_sample_document_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_sample_document_comment != '') or \
                    (obj.advertising_structures_has_document_template_comment != 'Да' and
                     obj.advertising_structures_has_document_template_comment != 'Не предусмотрено' and
                     obj.advertising_structures_has_document_template_comment != ''):
                objects['advertising_structures'].append(obj)

            if (obj.capital_construction_has_advanced_appointment_comment != 'Да' and
                obj.capital_construction_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.capital_construction_has_advanced_appointment_comment != '') or \
                (obj.capital_construction_has_btn_get_service_comment != 'Да' and
                 obj.capital_construction_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.capital_construction_has_btn_get_service_comment != '') or \
                    (obj.capital_construction_has_reglament_comment != 'Да' and
                     obj.capital_construction_has_reglament_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_reglament_comment != '') or \
                    (obj.capital_construction_has_estimation_quality_comment != 'Да' and
                     obj.capital_construction_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_estimation_quality_comment != '') or \
                    (obj.capital_construction_connected_to_fgis_do_comment != 'Да' and
                     obj.capital_construction_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.capital_construction_connected_to_fgis_do_comment != '') or \
                    (obj.capital_construction_has_electronic_form_printing_comment != 'Да' and
                     obj.capital_construction_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_electronic_form_printing_comment != '') or \
                    (obj.capital_construction_has_edition_draft_comment != 'Да' and
                     obj.capital_construction_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_edition_draft_comment != '') or \
                    (obj.capital_construction_has_term_of_consideration_comment != 'Да' and
                     obj.capital_construction_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_term_of_consideration_comment != '') or \
                    (obj.capital_construction_has_notif_consider_result_comment != 'Да' and
                     obj.capital_construction_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_notif_consider_result_comment != '') or \
                    (obj.capital_construction_has_causes_of_failure_comment != 'Да' and
                     obj.capital_construction_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_causes_of_failure_comment != '') or \
                    (obj.capital_construction_has_sample_document_comment != 'Да' and
                     obj.capital_construction_has_sample_document_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_sample_document_comment != '') or \
                    (obj.capital_construction_has_document_template_comment != 'Да' and
                     obj.capital_construction_has_document_template_comment != 'Не предусмотрено' and
                     obj.capital_construction_has_document_template_comment != ''):
                objects['capital_construction'].append(obj)

            if (obj.preschool_education_has_advanced_appointment_comment != 'Да' and
                obj.preschool_education_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.preschool_education_has_advanced_appointment_comment != '') or \
                (obj.preschool_education_has_btn_get_service_comment != 'Да' and
                 obj.preschool_education_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.preschool_education_has_btn_get_service_comment != '') or \
                    (obj.preschool_education_has_reglament_comment != 'Да' and
                     obj.preschool_education_has_reglament_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_reglament_comment != '') or \
                    (obj.preschool_education_has_estimation_quality_comment != 'Да' and
                     obj.preschool_education_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_estimation_quality_comment != '') or \
                    (obj.preschool_education_connected_to_fgis_do_comment != 'Да' and
                     obj.preschool_education_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.preschool_education_connected_to_fgis_do_comment != '') or \
                    (obj.preschool_education_has_electronic_form_printing_comment != 'Да' and
                     obj.preschool_education_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_electronic_form_printing_comment != '') or \
                    (obj.preschool_education_has_edition_draft_comment != 'Да' and
                     obj.preschool_education_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_edition_draft_comment != '') or \
                    (obj.preschool_education_has_term_of_consideration_comment != 'Да' and
                     obj.preschool_education_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_term_of_consideration_comment != '') or \
                    (obj.preschool_education_has_notif_consider_result_comment != 'Да' and
                     obj.preschool_education_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_notif_consider_result_comment != '') or \
                    (obj.preschool_education_has_causes_of_failure_comment != 'Да' and
                     obj.preschool_education_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_causes_of_failure_comment != '') or \
                    (obj.preschool_education_has_sample_document_comment != 'Да' and
                     obj.preschool_education_has_sample_document_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_sample_document_comment != '') or \
                    (obj.preschool_education_has_document_template_comment != 'Да' and
                     obj.preschool_education_has_document_template_comment != 'Не предусмотрено' and
                     obj.preschool_education_has_document_template_comment != ''):
                objects['preschool_education'].append(obj)

            if (obj.school_education_has_advanced_appointment_comment != 'Да' and
                obj.school_education_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.school_education_has_advanced_appointment_comment != '') or \
                (obj.school_education_has_btn_get_service_comment != 'Да' and
                 obj.school_education_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.school_education_has_btn_get_service_comment != '') or \
                    (obj.school_education_has_reglament_comment != 'Да' and
                     obj.school_education_has_reglament_comment != 'Не предусмотрено' and
                     obj.school_education_has_reglament_comment != '') or \
                    (obj.school_education_has_estimation_quality_comment != 'Да' and
                     obj.school_education_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.school_education_has_estimation_quality_comment != '') or \
                    (obj.school_education_connected_to_fgis_do_comment != 'Да' and
                     obj.school_education_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.school_education_connected_to_fgis_do_comment != '') or \
                    (obj.school_education_has_electronic_form_printing_comment != 'Да' and
                     obj.school_education_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.school_education_has_electronic_form_printing_comment != '') or \
                    (obj.school_education_has_edition_draft_comment != 'Да' and
                     obj.school_education_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.school_education_has_edition_draft_comment != '') or \
                    (obj.school_education_has_term_of_consideration_comment != 'Да' and
                     obj.school_education_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.school_education_has_term_of_consideration_comment != '') or \
                    (obj.school_education_has_notif_consider_result_comment != 'Да' and
                     obj.school_education_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.school_education_has_notif_consider_result_comment != '') or \
                    (obj.school_education_has_causes_of_failure_comment != 'Да' and
                     obj.school_education_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.school_education_has_causes_of_failure_comment != '') or \
                    (obj.school_education_has_sample_document_comment != 'Да' and
                     obj.school_education_has_sample_document_comment != 'Не предусмотрено' and
                     obj.school_education_has_sample_document_comment != '') or \
                    (obj.school_education_has_document_template_comment != 'Да' and
                     obj.school_education_has_document_template_comment != 'Не предусмотрено' and
                     obj.school_education_has_document_template_comment != ''):
                objects['school_education'].append(obj)

            if (obj.needing_premises_has_advanced_appointment_comment != 'Да' and
                obj.needing_premises_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.needing_premises_has_advanced_appointment_comment != '') or \
                (obj.needing_premises_has_btn_get_service_comment != 'Да' and
                 obj.needing_premises_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.needing_premises_has_btn_get_service_comment != '') or \
                    (obj.needing_premises_has_reglament_comment != 'Да' and
                     obj.needing_premises_has_reglament_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_reglament_comment != '') or \
                    (obj.needing_premises_has_estimation_quality_comment != 'Да' and
                     obj.needing_premises_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_estimation_quality_comment != '') or \
                    (obj.needing_premises_connected_to_fgis_do_comment != 'Да' and
                     obj.needing_premises_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.needing_premises_connected_to_fgis_do_comment != '') or \
                    (obj.needing_premises_has_electronic_form_printing_comment != 'Да' and
                     obj.needing_premises_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_electronic_form_printing_comment != '') or \
                    (obj.needing_premises_has_edition_draft_comment != 'Да' and
                     obj.needing_premises_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_edition_draft_comment != '') or \
                    (obj.needing_premises_has_term_of_consideration_comment != 'Да' and
                     obj.needing_premises_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_term_of_consideration_comment != '') or \
                    (obj.needing_premises_has_notif_consider_result_comment != 'Да' and
                     obj.needing_premises_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_notif_consider_result_comment != '') or \
                    (obj.needing_premises_has_causes_of_failure_comment != 'Да' and
                     obj.needing_premises_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_causes_of_failure_comment != '') or \
                    (obj.needing_premises_has_sample_document_comment != 'Да' and
                     obj.needing_premises_has_sample_document_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_sample_document_comment != '') or \
                    (obj.needing_premises_has_document_template_comment != 'Да' and
                     obj.needing_premises_has_document_template_comment != 'Не предусмотрено' and
                     obj.needing_premises_has_document_template_comment != ''):
                objects['needing_premises'].append(obj)

            if (obj.town_planning_has_advanced_appointment_comment != 'Да' and
                obj.town_planning_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.town_planning_has_advanced_appointment_comment != '') or \
                (obj.town_planning_has_btn_get_service_comment != 'Да' and
                 obj.town_planning_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.town_planning_has_btn_get_service_comment != '') or \
                    (obj.town_planning_has_reglament_comment != 'Да' and
                     obj.town_planning_has_reglament_comment != 'Не предусмотрено' and
                     obj.town_planning_has_reglament_comment != '') or \
                    (obj.town_planning_has_estimation_quality_comment != 'Да' and
                     obj.town_planning_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.town_planning_has_estimation_quality_comment != '') or \
                    (obj.town_planning_connected_to_fgis_do_comment != 'Да' and
                     obj.town_planning_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.town_planning_connected_to_fgis_do_comment != '') or \
                    (obj.town_planning_has_electronic_form_printing_comment != 'Да' and
                     obj.town_planning_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.town_planning_has_electronic_form_printing_comment != '') or \
                    (obj.town_planning_has_edition_draft_comment != 'Да' and
                     obj.town_planning_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.town_planning_has_edition_draft_comment != '') or \
                    (obj.town_planning_has_term_of_consideration_comment != 'Да' and
                     obj.town_planning_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.town_planning_has_term_of_consideration_comment != '') or \
                    (obj.town_planning_has_notif_consider_result_comment != 'Да' and
                     obj.town_planning_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.town_planning_has_notif_consider_result_comment != '') or \
                    (obj.town_planning_has_causes_of_failure_comment != 'Да' and
                     obj.town_planning_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.town_planning_has_causes_of_failure_comment != '') or \
                    (obj.town_planning_has_sample_document_comment != 'Да' and
                     obj.town_planning_has_sample_document_comment != 'Не предусмотрено' and
                     obj.town_planning_has_sample_document_comment != '') or \
                    (obj.town_planning_has_document_template_comment != 'Да' and
                     obj.town_planning_has_document_template_comment != 'Не предусмотрено' and
                     obj.town_planning_has_document_template_comment != ''):
                objects['town_planning'].append(obj)

            if (obj.archive_reference_has_advanced_appointment_comment != 'Да' and
                obj.archive_reference_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.archive_reference_has_advanced_appointment_comment != '') or \
                (obj.archive_reference_has_btn_get_service_comment != 'Да' and
                 obj.archive_reference_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.archive_reference_has_btn_get_service_comment != '') or \
                    (obj.archive_reference_has_reglament_comment != 'Да' and
                     obj.archive_reference_has_reglament_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_reglament_comment != '') or \
                    (obj.archive_reference_has_estimation_quality_comment != 'Да' and
                     obj.archive_reference_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_estimation_quality_comment != '') or \
                    (obj.archive_reference_connected_to_fgis_do_comment != 'Да' and
                     obj.archive_reference_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.archive_reference_connected_to_fgis_do_comment != '') or \
                    (obj.archive_reference_has_electronic_form_printing_comment != 'Да' and
                     obj.archive_reference_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_electronic_form_printing_comment != '') or \
                    (obj.archive_reference_has_edition_draft_comment != 'Да' and
                     obj.archive_reference_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_edition_draft_comment != '') or \
                    (obj.archive_reference_has_term_of_consideration_comment != 'Да' and
                     obj.archive_reference_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_term_of_consideration_comment != '') or \
                    (obj.archive_reference_has_notif_consider_result_comment != 'Да' and
                     obj.archive_reference_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_notif_consider_result_comment != '') or \
                    (obj.archive_reference_has_causes_of_failure_comment != 'Да' and
                     obj.archive_reference_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_causes_of_failure_comment != '') or \
                    (obj.archive_reference_has_sample_document_comment != 'Да' and
                     obj.archive_reference_has_sample_document_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_sample_document_comment != '') or \
                    (obj.archive_reference_has_document_template_comment != 'Да' and
                     obj.archive_reference_has_document_template_comment != 'Не предусмотрено' and
                     obj.archive_reference_has_document_template_comment != ''):
                objects['archive_reference'].append(obj)

            if (obj.land_schemes_has_advanced_appointment_comment != 'Да' and
                obj.land_schemes_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.land_schemes_has_advanced_appointment_comment != '') or \
                (obj.land_schemes_has_btn_get_service_comment != 'Да' and
                 obj.land_schemes_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.land_schemes_has_btn_get_service_comment != '') or \
                    (obj.land_schemes_has_reglament_comment != 'Да' and
                     obj.land_schemes_has_reglament_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_reglament_comment != '') or \
                    (obj.land_schemes_has_estimation_quality_comment != 'Да' and
                     obj.land_schemes_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_estimation_quality_comment != '') or \
                    (obj.land_schemes_connected_to_fgis_do_comment != 'Да' and
                     obj.land_schemes_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.land_schemes_connected_to_fgis_do_comment != '') or \
                    (obj.land_schemes_has_electronic_form_printing_comment != 'Да' and
                     obj.land_schemes_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_electronic_form_printing_comment != '') or \
                    (obj.land_schemes_has_edition_draft_comment != 'Да' and
                     obj.land_schemes_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_edition_draft_comment != '') or \
                    (obj.land_schemes_has_term_of_consideration_comment != 'Да' and
                     obj.land_schemes_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_term_of_consideration_comment != '') or \
                    (obj.land_schemes_has_notif_consider_result_comment != 'Да' and
                     obj.land_schemes_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_notif_consider_result_comment != '') or \
                    (obj.land_schemes_has_causes_of_failure_comment != 'Да' and
                     obj.land_schemes_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_causes_of_failure_comment != '') or \
                    (obj.land_schemes_has_sample_document_comment != 'Да' and
                     obj.land_schemes_has_sample_document_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_sample_document_comment != '') or \
                    (obj.land_schemes_has_document_template_comment != 'Да' and
                     obj.land_schemes_has_document_template_comment != 'Не предусмотрено' and
                     obj.land_schemes_has_document_template_comment != ''):
                objects['land_schemes'].append(obj)

            if (obj.land_sale_has_advanced_appointment_comment != 'Да' and
                obj.land_sale_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.land_sale_has_advanced_appointment_comment != '') or \
                (obj.land_sale_has_btn_get_service_comment != 'Да' and
                 obj.land_sale_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.land_sale_has_btn_get_service_comment != '') or \
                    (obj.land_sale_has_reglament_comment != 'Да' and
                     obj.land_sale_has_reglament_comment != 'Не предусмотрено' and
                     obj.land_sale_has_reglament_comment != '') or \
                    (obj.land_sale_has_estimation_quality_comment != 'Да' and
                     obj.land_sale_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.land_sale_has_estimation_quality_comment != '') or \
                    (obj.land_sale_connected_to_fgis_do_comment != 'Да' and
                     obj.land_sale_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.land_sale_connected_to_fgis_do_comment != '') or \
                    (obj.land_sale_has_electronic_form_printing_comment != 'Да' and
                     obj.land_sale_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.land_sale_has_electronic_form_printing_comment != '') or \
                    (obj.land_sale_has_edition_draft_comment != 'Да' and
                     obj.land_sale_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.land_sale_has_edition_draft_comment != '') or \
                    (obj.land_sale_has_term_of_consideration_comment != 'Да' and
                     obj.land_sale_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.land_sale_has_term_of_consideration_comment != '') or \
                    (obj.land_sale_has_notif_consider_result_comment != 'Да' and
                     obj.land_sale_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.land_sale_has_notif_consider_result_comment != '') or \
                    (obj.land_sale_has_causes_of_failure_comment != 'Да' and
                     obj.land_sale_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.land_sale_has_causes_of_failure_comment != '') or \
                    (obj.land_sale_has_sample_document_comment != 'Да' and
                     obj.land_sale_has_sample_document_comment != 'Не предусмотрено' and
                     obj.land_sale_has_sample_document_comment != '') or \
                    (obj.land_sale_has_document_template_comment != 'Да' and
                     obj.land_sale_has_document_template_comment != 'Не предусмотрено' and
                     obj.land_sale_has_document_template_comment != ''):
                objects['land_sale'].append(obj)

            if (obj.land_lease_has_advanced_appointment_comment != 'Да' and
                obj.land_lease_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.land_lease_has_advanced_appointment_comment != '') or \
                (obj.land_lease_has_btn_get_service_comment != 'Да' and
                 obj.land_lease_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.land_lease_has_btn_get_service_comment != '') or \
                    (obj.land_lease_has_reglament_comment != 'Да' and
                     obj.land_lease_has_reglament_comment != 'Не предусмотрено' and
                     obj.land_lease_has_reglament_comment != '') or \
                    (obj.land_lease_has_estimation_quality_comment != 'Да' and
                     obj.land_lease_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.land_lease_has_estimation_quality_comment != '') or \
                    (obj.land_lease_connected_to_fgis_do_comment != 'Да' and
                     obj.land_lease_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.land_lease_connected_to_fgis_do_comment != '') or \
                    (obj.land_lease_has_electronic_form_printing_comment != 'Да' and
                     obj.land_lease_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.land_lease_has_electronic_form_printing_comment != '') or \
                    (obj.land_lease_has_edition_draft_comment != 'Да' and
                     obj.land_lease_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.land_lease_has_edition_draft_comment != '') or \
                    (obj.land_lease_has_term_of_consideration_comment != 'Да' and
                     obj.land_lease_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.land_lease_has_term_of_consideration_comment != '') or \
                    (obj.land_lease_has_notif_consider_result_comment != 'Да' and
                     obj.land_lease_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.land_lease_has_notif_consider_result_comment != '') or \
                    (obj.land_lease_has_causes_of_failure_comment != 'Да' and
                     obj.land_lease_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.land_lease_has_causes_of_failure_comment != '') or \
                    (obj.land_lease_has_sample_document_comment != 'Да' and
                     obj.land_lease_has_sample_document_comment != 'Не предусмотрено' and
                     obj.land_lease_has_sample_document_comment != '') or \
                    (obj.land_lease_has_document_template_comment != 'Да' and
                     obj.land_lease_has_document_template_comment != 'Не предусмотрено' and
                     obj.land_lease_has_document_template_comment != ''):
                objects['land_lease'].append(obj)

            if (obj.ownership_right_has_advanced_appointment_comment != 'Да' and
                obj.ownership_right_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.ownership_right_has_advanced_appointment_comment != '') or \
                (obj.ownership_right_has_btn_get_service_comment != 'Да' and
                 obj.ownership_right_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.ownership_right_has_btn_get_service_comment != '') or \
                    (obj.ownership_right_has_reglament_comment != 'Да' and
                     obj.ownership_right_has_reglament_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_reglament_comment != '') or \
                    (obj.ownership_right_has_estimation_quality_comment != 'Да' and
                     obj.ownership_right_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_estimation_quality_comment != '') or \
                    (obj.ownership_right_connected_to_fgis_do_comment != 'Да' and
                     obj.ownership_right_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.ownership_right_connected_to_fgis_do_comment != '') or \
                    (obj.ownership_right_has_electronic_form_printing_comment != 'Да' and
                     obj.ownership_right_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_electronic_form_printing_comment != '') or \
                    (obj.ownership_right_has_edition_draft_comment != 'Да' and
                     obj.ownership_right_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_edition_draft_comment != '') or \
                    (obj.ownership_right_has_term_of_consideration_comment != 'Да' and
                     obj.ownership_right_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_term_of_consideration_comment != '') or \
                    (obj.ownership_right_has_notif_consider_result_comment != 'Да' and
                     obj.ownership_right_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_notif_consider_result_comment != '') or \
                    (obj.ownership_right_has_causes_of_failure_comment != 'Да' and
                     obj.ownership_right_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_causes_of_failure_comment != '') or \
                    (obj.ownership_right_has_sample_document_comment != 'Да' and
                     obj.ownership_right_has_sample_document_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_sample_document_comment != '') or \
                    (obj.ownership_right_has_document_template_comment != 'Да' and
                     obj.ownership_right_has_document_template_comment != 'Не предусмотрено' and
                     obj.ownership_right_has_document_template_comment != ''):
                objects['ownership_right'].append(obj)

            if (obj.municipal_property_lease_has_advanced_appointment_comment != 'Да' and
                obj.municipal_property_lease_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.municipal_property_lease_has_advanced_appointment_comment != '') or \
                (obj.municipal_property_lease_has_btn_get_service_comment != 'Да' and
                 obj.municipal_property_lease_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.municipal_property_lease_has_btn_get_service_comment != '') or \
                    (obj.municipal_property_lease_has_reglament_comment != 'Да' and
                     obj.municipal_property_lease_has_reglament_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_reglament_comment != '') or \
                    (obj.municipal_property_lease_has_estimation_quality_comment != 'Да' and
                     obj.municipal_property_lease_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_estimation_quality_comment != '') or \
                    (obj.municipal_property_lease_connected_to_fgis_do_comment != 'Да' and
                     obj.municipal_property_lease_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_connected_to_fgis_do_comment != '') or \
                    (obj.municipal_property_lease_has_electronic_form_printing_comment != 'Да' and
                     obj.municipal_property_lease_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_electronic_form_printing_comment != '') or \
                    (obj.municipal_property_lease_has_edition_draft_comment != 'Да' and
                     obj.municipal_property_lease_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_edition_draft_comment != '') or \
                    (obj.municipal_property_lease_has_term_of_consideration_comment != 'Да' and
                     obj.municipal_property_lease_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_term_of_consideration_comment != '') or \
                    (obj.municipal_property_lease_has_notif_consider_result_comment != 'Да' and
                     obj.municipal_property_lease_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_notif_consider_result_comment != '') or \
                    (obj.municipal_property_lease_has_causes_of_failure_comment != 'Да' and
                     obj.municipal_property_lease_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_causes_of_failure_comment != '') or \
                    (obj.municipal_property_lease_has_sample_document_comment != 'Да' and
                     obj.municipal_property_lease_has_sample_document_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_sample_document_comment != '') or \
                    (obj.municipal_property_lease_has_document_template_comment != 'Да' and
                     obj.municipal_property_lease_has_document_template_comment != 'Не предусмотрено' and
                     obj.municipal_property_lease_has_document_template_comment != ''):
                objects['municipal_property_lease'].append(obj)

            if (obj.free_land_provision_has_advanced_appointment_comment != 'Да' and
                obj.free_land_provision_has_advanced_appointment_comment != 'Не предусмотрено' and
                obj.free_land_provision_has_advanced_appointment_comment != '') or \
                (obj.free_land_provision_has_btn_get_service_comment != 'Да' and
                 obj.free_land_provision_has_btn_get_service_comment != 'Не предусмотрено' and
                 obj.free_land_provision_has_btn_get_service_comment != '') or \
                    (obj.free_land_provision_has_reglament_comment != 'Да' and
                     obj.free_land_provision_has_reglament_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_reglament_comment != '') or \
                    (obj.free_land_provision_has_estimation_quality_comment != 'Да' and
                     obj.free_land_provision_has_estimation_quality_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_estimation_quality_comment != '') or \
                    (obj.free_land_provision_connected_to_fgis_do_comment != 'Да' and
                     obj.free_land_provision_connected_to_fgis_do_comment != 'Не предусмотрено' and
                     obj.free_land_provision_connected_to_fgis_do_comment != '') or \
                    (obj.free_land_provision_has_electronic_form_printing_comment != 'Да' and
                     obj.free_land_provision_has_electronic_form_printing_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_electronic_form_printing_comment != '') or \
                    (obj.free_land_provision_has_edition_draft_comment != 'Да' and
                     obj.free_land_provision_has_edition_draft_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_edition_draft_comment != '') or \
                    (obj.free_land_provision_has_term_of_consideration_comment != 'Да' and
                     obj.free_land_provision_has_term_of_consideration_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_term_of_consideration_comment != '') or \
                    (obj.free_land_provision_has_notif_consider_result_comment != 'Да' and
                     obj.free_land_provision_has_notif_consider_result_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_notif_consider_result_comment != '') or \
                    (obj.free_land_provision_has_causes_of_failure_comment != 'Да' and
                     obj.free_land_provision_has_causes_of_failure_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_causes_of_failure_comment != '') or \
                    (obj.free_land_provision_has_sample_document_comment != 'Да' and
                     obj.free_land_provision_has_sample_document_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_sample_document_comment != '') or \
                    (obj.free_land_provision_has_document_template_comment != 'Да' and
                     obj.free_land_provision_has_document_template_comment != 'Не предусмотрено' and
                     obj.free_land_provision_has_document_template_comment != ''):
                objects['free_land_provision'].append(obj)
        except Http404:
            pass
    return objects

@login_required(login_url='/login/',
                redirect_field_name='/result_form/with_no_troubles/' + datetime.today().strftime('%Y/%m/'))
def get_with_no_troubles(request, month, year):
    region_full_names = []
    for name in regions_names:
        try:
            obj = get_object_or_404(RegionModel, region_name=name, month=month, year=year)
            if (obj.residential_premises_has_advanced_appointment_comment == 'Да' or
                obj.residential_premises_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_btn_get_service_comment == 'Да' or
                     obj.residential_premises_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_reglament_comment == 'Да' or
                     obj.residential_premises_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_estimation_quality_comment == 'Да' or
                     obj.residential_premises_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_connected_to_fgis_do_comment == 'Да' or
                     obj.residential_premises_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_electronic_form_printing_comment == 'Да' or
                     obj.residential_premises_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_edition_draft_comment == 'Да' or
                     obj.residential_premises_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_term_of_consideration_comment == 'Да' or
                     obj.residential_premises_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_notif_consider_result_comment == 'Да' or
                     obj.residential_premises_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_causes_of_failure_comment == 'Да' or
                     obj.residential_premises_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_sample_document_comment == 'Да' or
                     obj.residential_premises_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.residential_premises_has_document_template_comment == 'Да' or
                     obj.residential_premises_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_advanced_appointment_comment == 'Да' or
                    obj.housing_transfer_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_btn_get_service_comment == 'Да' or
                     obj.housing_transfer_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_reglament_comment == 'Да' or
                     obj.housing_transfer_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_estimation_quality_comment == 'Да' or
                     obj.housing_transfer_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_connected_to_fgis_do_comment == 'Да' or
                     obj.housing_transfer_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_electronic_form_printing_comment == 'Да' or
                     obj.housing_transfer_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_edition_draft_comment == 'Да' or
                     obj.housing_transfer_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_term_of_consideration_comment == 'Да' or
                     obj.housing_transfer_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_notif_consider_result_comment == 'Да' or
                     obj.housing_transfer_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_causes_of_failure_comment == 'Да' or
                     obj.housing_transfer_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_sample_document_comment == 'Да' or
                     obj.housing_transfer_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.housing_transfer_has_document_template_comment == 'Да' or
                     obj.housing_transfer_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_advanced_appointment_comment == 'Да' or
                    obj.advertising_structures_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_btn_get_service_comment == 'Да' or
                     obj.advertising_structures_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_reglament_comment == 'Да' or
                     obj.advertising_structures_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_estimation_quality_comment == 'Да' or
                     obj.advertising_structures_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_connected_to_fgis_do_comment == 'Да' or
                     obj.advertising_structures_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_electronic_form_printing_comment == 'Да' or
                     obj.advertising_structures_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_edition_draft_comment == 'Да' or
                     obj.advertising_structures_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_term_of_consideration_comment == 'Да' or
                     obj.advertising_structures_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_notif_consider_result_comment == 'Да' or
                     obj.advertising_structures_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_causes_of_failure_comment == 'Да' or
                     obj.advertising_structures_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_sample_document_comment == 'Да' or
                     obj.advertising_structures_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.advertising_structures_has_document_template_comment == 'Да' or
                     obj.advertising_structures_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_advanced_appointment_comment == 'Да' or
                    obj.capital_construction_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_btn_get_service_comment == 'Да' or
                     obj.capital_construction_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_reglament_comment == 'Да' or
                     obj.capital_construction_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_estimation_quality_comment == 'Да' or
                     obj.capital_construction_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_connected_to_fgis_do_comment == 'Да' or
                     obj.capital_construction_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_electronic_form_printing_comment == 'Да' or
                     obj.capital_construction_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_edition_draft_comment == 'Да' or
                     obj.capital_construction_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_term_of_consideration_comment == 'Да' or
                     obj.capital_construction_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_notif_consider_result_comment == 'Да' or
                     obj.capital_construction_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_causes_of_failure_comment == 'Да' or
                     obj.capital_construction_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_sample_document_comment == 'Да' or
                     obj.capital_construction_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.capital_construction_has_document_template_comment == 'Да' or
                     obj.capital_construction_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_advanced_appointment_comment == 'Да' or
                    obj.preschool_education_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_btn_get_service_comment == 'Да' or
                     obj.preschool_education_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_reglament_comment == 'Да' or
                     obj.preschool_education_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_estimation_quality_comment == 'Да' or
                     obj.preschool_education_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_connected_to_fgis_do_comment == 'Да' or
                     obj.preschool_education_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_electronic_form_printing_comment == 'Да' or
                     obj.preschool_education_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_edition_draft_comment == 'Да' or
                     obj.preschool_education_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_term_of_consideration_comment == 'Да' or
                     obj.preschool_education_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_notif_consider_result_comment == 'Да' or
                     obj.preschool_education_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_causes_of_failure_comment == 'Да' or
                     obj.preschool_education_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_sample_document_comment == 'Да' or
                     obj.preschool_education_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.preschool_education_has_document_template_comment == 'Да' or
                     obj.preschool_education_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_advanced_appointment_comment == 'Да' or
                    obj.school_education_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_btn_get_service_comment == 'Да' or
                     obj.school_education_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_reglament_comment == 'Да' or
                     obj.school_education_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_estimation_quality_comment == 'Да' or
                     obj.school_education_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.school_education_connected_to_fgis_do_comment == 'Да' or
                     obj.school_education_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_electronic_form_printing_comment == 'Да' or
                     obj.school_education_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_edition_draft_comment == 'Да' or
                     obj.school_education_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_term_of_consideration_comment == 'Да' or
                     obj.school_education_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_notif_consider_result_comment == 'Да' or
                     obj.school_education_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_causes_of_failure_comment == 'Да' or
                     obj.school_education_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_sample_document_comment == 'Да' or
                     obj.school_education_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.school_education_has_document_template_comment == 'Да' or
                     obj.school_education_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_advanced_appointment_comment == 'Да' or
                    obj.needing_premises_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_btn_get_service_comment == 'Да' or
                     obj.needing_premises_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_reglament_comment == 'Да' or
                     obj.needing_premises_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_estimation_quality_comment == 'Да' or
                     obj.needing_premises_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_connected_to_fgis_do_comment == 'Да' or
                     obj.needing_premises_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_electronic_form_printing_comment == 'Да' or
                     obj.needing_premises_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_edition_draft_comment == 'Да' or
                     obj.needing_premises_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_term_of_consideration_comment == 'Да' or
                     obj.needing_premises_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_notif_consider_result_comment == 'Да' or
                     obj.needing_premises_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_causes_of_failure_comment == 'Да' or
                     obj.needing_premises_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_sample_document_comment == 'Да' or
                     obj.needing_premises_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.needing_premises_has_document_template_comment == 'Да' or
                     obj.needing_premises_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_advanced_appointment_comment == 'Да' or
                    obj.town_planning_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_btn_get_service_comment == 'Да' or
                     obj.town_planning_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_reglament_comment == 'Да' or
                     obj.town_planning_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_estimation_quality_comment == 'Да' or
                     obj.town_planning_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.town_planning_connected_to_fgis_do_comment == 'Да' or
                     obj.town_planning_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_electronic_form_printing_comment == 'Да' or
                     obj.town_planning_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_edition_draft_comment == 'Да' or
                     obj.town_planning_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_term_of_consideration_comment == 'Да' or
                     obj.town_planning_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_notif_consider_result_comment == 'Да' or
                     obj.town_planning_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_causes_of_failure_comment == 'Да' or
                     obj.town_planning_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_sample_document_comment == 'Да' or
                     obj.town_planning_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.town_planning_has_document_template_comment == 'Да' or
                     obj.town_planning_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_advanced_appointment_comment == 'Да' or
                    obj.archive_reference_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_btn_get_service_comment == 'Да' or
                     obj.archive_reference_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_reglament_comment == 'Да' or
                     obj.archive_reference_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_estimation_quality_comment == 'Да' or
                     obj.archive_reference_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_connected_to_fgis_do_comment == 'Да' or
                     obj.archive_reference_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_electronic_form_printing_comment == 'Да' or
                     obj.archive_reference_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_edition_draft_comment == 'Да' or
                     obj.archive_reference_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_term_of_consideration_comment == 'Да' or
                     obj.archive_reference_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_notif_consider_result_comment == 'Да' or
                     obj.archive_reference_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_causes_of_failure_comment == 'Да' or
                     obj.archive_reference_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_sample_document_comment == 'Да' or
                     obj.archive_reference_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.archive_reference_has_document_template_comment == 'Да' or
                     obj.archive_reference_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_advanced_appointment_comment == 'Да' or
                    obj.land_schemes_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_btn_get_service_comment == 'Да' or
                     obj.land_schemes_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_reglament_comment == 'Да' or
                     obj.land_schemes_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_estimation_quality_comment == 'Да' or
                     obj.land_schemes_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_connected_to_fgis_do_comment == 'Да' or
                     obj.land_schemes_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_electronic_form_printing_comment == 'Да' or
                     obj.land_schemes_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_edition_draft_comment == 'Да' or
                     obj.land_schemes_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_term_of_consideration_comment == 'Да' or
                     obj.land_schemes_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_notif_consider_result_comment == 'Да' or
                     obj.land_schemes_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_causes_of_failure_comment == 'Да' or
                     obj.land_schemes_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_sample_document_comment == 'Да' or
                     obj.land_schemes_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.land_schemes_has_document_template_comment == 'Да' or
                     obj.land_schemes_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_advanced_appointment_comment == 'Да' or
                     obj.land_sale_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_btn_get_service_comment == 'Да' or
                     obj.land_sale_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_reglament_comment == 'Да' or
                     obj.land_sale_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_estimation_quality_comment == 'Да' or
                     obj.land_sale_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.land_sale_connected_to_fgis_do_comment == 'Да' or
                     obj.land_sale_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_electronic_form_printing_comment == 'Да' or
                     obj.land_sale_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_edition_draft_comment == 'Да' or
                     obj.land_sale_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_term_of_consideration_comment == 'Да' or
                     obj.land_sale_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_notif_consider_result_comment == 'Да' or
                     obj.land_sale_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_causes_of_failure_comment == 'Да' or
                     obj.land_sale_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_sample_document_comment == 'Да' or
                     obj.land_sale_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.land_sale_has_document_template_comment == 'Да' or
                     obj.land_sale_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_advanced_appointment_comment == 'Да' or
                     obj.land_lease_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_btn_get_service_comment == 'Да' or
                     obj.land_lease_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_reglament_comment == 'Да' or
                     obj.land_lease_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_estimation_quality_comment == 'Да' or
                     obj.land_lease_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.land_lease_connected_to_fgis_do_comment == 'Да' or
                     obj.land_lease_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_electronic_form_printing_comment == 'Да' or
                     obj.land_lease_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_edition_draft_comment == 'Да' or
                     obj.land_lease_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_term_of_consideration_comment == 'Да' or
                     obj.land_lease_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_notif_consider_result_comment == 'Да' or
                     obj.land_lease_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_causes_of_failure_comment == 'Да' or
                     obj.land_lease_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_sample_document_comment == 'Да' or
                     obj.land_lease_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.land_lease_has_document_template_comment == 'Да' or
                     obj.land_lease_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_advanced_appointment_comment == 'Да' or
                     obj.ownership_right_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_btn_get_service_comment == 'Да' or
                     obj.ownership_right_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_reglament_comment == 'Да' or
                     obj.ownership_right_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_estimation_quality_comment == 'Да' or
                     obj.ownership_right_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_connected_to_fgis_do_comment == 'Да' or
                     obj.ownership_right_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_electronic_form_printing_comment == 'Да' or
                     obj.ownership_right_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_edition_draft_comment == 'Да' or
                     obj.ownership_right_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_term_of_consideration_comment == 'Да' or
                     obj.ownership_right_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_notif_consider_result_comment == 'Да' or
                     obj.ownership_right_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_causes_of_failure_comment == 'Да' or
                     obj.ownership_right_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_sample_document_comment == 'Да' or
                     obj.ownership_right_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.ownership_right_has_document_template_comment == 'Да' or
                     obj.ownership_right_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_advanced_appointment_comment == 'Да' or
                     obj.municipal_property_lease_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_btn_get_service_comment == 'Да' or
                     obj.municipal_property_lease_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_reglament_comment == 'Да' or
                     obj.municipal_property_lease_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_estimation_quality_comment == 'Да' or
                     obj.municipal_property_lease_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_connected_to_fgis_do_comment == 'Да' or
                     obj.municipal_property_lease_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_electronic_form_printing_comment == 'Да' or
                     obj.municipal_property_lease_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_edition_draft_comment == 'Да' or
                     obj.municipal_property_lease_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_term_of_consideration_comment == 'Да' or
                     obj.municipal_property_lease_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_notif_consider_result_comment == 'Да' or
                     obj.municipal_property_lease_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_causes_of_failure_comment == 'Да' or
                     obj.municipal_property_lease_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_sample_document_comment == 'Да' or
                     obj.municipal_property_lease_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.municipal_property_lease_has_document_template_comment == 'Да' or
                     obj.municipal_property_lease_has_document_template_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_advanced_appointment_comment == 'Да' or
                     obj.free_land_provision_has_advanced_appointment_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_btn_get_service_comment == 'Да' or
                     obj.free_land_provision_has_btn_get_service_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_reglament_comment == 'Да' or
                     obj.free_land_provision_has_reglament_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_estimation_quality_comment == 'Да' or
                     obj.free_land_provision_has_estimation_quality_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_connected_to_fgis_do_comment == 'Да' or
                     obj.free_land_provision_connected_to_fgis_do_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_electronic_form_printing_comment == 'Да' or
                     obj.free_land_provision_has_electronic_form_printing_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_edition_draft_comment == 'Да' or
                     obj.free_land_provision_has_edition_draft_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_term_of_consideration_comment == 'Да' or
                     obj.free_land_provision_has_term_of_consideration_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_notif_consider_result_comment == 'Да' or
                     obj.free_land_provision_has_notif_consider_result_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_causes_of_failure_comment == 'Да' or
                     obj.free_land_provision_has_causes_of_failure_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_sample_document_comment == 'Да' or
                     obj.free_land_provision_has_sample_document_comment == 'Не предусмотрено') and \
                    (obj.free_land_provision_has_document_template_comment == 'Да' or
                     obj.free_land_provision_has_document_template_comment == 'Не предусмотрено'):
                region_full_names.append(obj.region_name)
        except Http404:
            pass
    indices = [regions_names.index(full_name) for full_name in region_full_names]
    return render(request, 'app/with_no_troubles.html', {
                                                    'zipped_names': zip(region_full_names, [short_regions_names[i] for i in indices]),
                                                     'year': str(year),
                                                     'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                     'num_month': month,
                                                     'zipped': zip(regions_names, short_regions_names),
                                                     'username': request.user.username,
                                                     'years': [i for i in range(2016, int(datetime.now().year) + 1)],
                                                    'zipped_service_names': zip(full_service_names, short_service_names)
                                                     })

@login_required(login_url='/login/', redirect_field_name='/form/')
def get_self_examination_form(request):
    """
    View function for renewing a specific SE_Form by users
    """
    if request.user.is_authenticated:

        if request.method == 'POST':
            region_form = SE_Form(request.POST)
            try:
                #trying to find existing form of current month to fill it with new information
                instance = get_object_or_404(RegionModel, region_name=request.user.region_name, month=datetime.today().strftime('%m'), year=datetime.today().strftime('%Y'))
                form_to_save = SE_Form(request.POST, instance=instance)
                if form_to_save.is_valid():
                    form_to_save.save(commit=False)
                    form_to_save.time = datetime.today().time()
                    form_to_save.day = datetime.today().strftime('%d')
                    form_to_save.save()
            except Http404:
                #else creating a new one, validationg and save

                if region_form.is_valid():
                    form_to_save = region_form.save(commit=False)
                    form_to_save.time = datetime.today().time()
                    form_to_save.day = datetime.today().strftime('%d')
                    form_to_save.year = datetime.today().strftime('%Y')
                    form_to_save.month = datetime.today().strftime('%m')
                    form_to_save.region_name = request.user.region_name
                    form_to_save.save()

            if request.POST['btn_action'] == 'exit':
                    # redirect to a new URL:
                    return HttpResponseRedirect(reverse('result_form', kwargs={
                    'service_name': 'residential_premises',
                    'year': datetime.today().strftime('%Y'),
                    'month': datetime.today().strftime('%m')
                    }))


        # If this is a GET (or any other method) create the default form.
        else:
            try:
                #trying to find existing form of current month for further user editition
                instance = get_object_or_404(RegionModel, region_name=request.user.region_name, month=datetime.today().strftime('%m'), year=datetime.today().strftime('%Y'))
                region_form = SE_Form(instance=instance)
            except Http404:
                #else creating a new empty one
                region_form = SE_Form()

        return render(request, 'app/form.html', {
            'region_form': region_form,
            'username': request.user.username,
            'region_name': request.user.region_name,
            'year':datetime.today().strftime('%Y'),
            'num_month':datetime.today().strftime('%m'),
            'zipped': zip(regions_names, short_regions_names),
            'full_service_names': full_service_names,
            'zipped_service_names': zip(full_service_names, short_service_names)
        })
    else:
        return HttpResponseRedirect(reverse('login'))


@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def get_result_form(request, service_name, year, month):
    objects = list()

    if service_name == 'residential_premises':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw('''SELECT id, region_name,
                    residential_premises_id_rgmu AS id_rgmu,
                    residential_premises_statement_amount AS statement_amount,
                    residential_premises_link AS link,
                    residential_premises_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    residential_premises_has_btn_get_service_comment AS has_btn_get_service_comment,
                    residential_premises_has_reglament_comment AS has_reglament_comment,
                    residential_premises_has_estimation_quality_comment AS has_estimation_quality_comment,
                    residential_premises_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    residential_premises_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    residential_premises_has_edition_draft_comment AS has_edition_draft_comment,
                    residential_premises_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    residential_premises_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    residential_premises_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    residential_premises_has_sample_document_comment AS has_sample_document_comment,
                    residential_premises_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                    month) + '''\' AND year=\'''' + str(
                    year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)

            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'housing_transfer':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    housing_transfer_id_rgmu AS id_rgmu,
                    housing_transfer_statement_amount AS statement_amount,
                    housing_transfer_link AS link,
                    housing_transfer_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    housing_transfer_has_btn_get_service_comment AS has_btn_get_service_comment,
                    housing_transfer_has_reglament_comment AS has_reglament_comment,
                    housing_transfer_has_estimation_quality_comment AS has_estimation_quality_comment,
                    housing_transfer_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    housing_transfer_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    housing_transfer_has_edition_draft_comment AS has_edition_draft_comment,
                    housing_transfer_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    housing_transfer_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    housing_transfer_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    housing_transfer_has_sample_document_comment AS has_sample_document_comment,
                    housing_transfer_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'advertising_structures':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    advertising_structures_id_rgmu AS id_rgmu,
                    advertising_structures_statement_amount AS statement_amount,
                    advertising_structures_link AS link,
                    advertising_structures_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    advertising_structures_has_btn_get_service_comment AS has_btn_get_service_comment,
                    advertising_structures_has_reglament_comment AS has_reglament_comment,
                    advertising_structures_has_estimation_quality_comment AS has_estimation_quality_comment,
                    advertising_structures_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    advertising_structures_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    advertising_structures_has_edition_draft_comment AS has_edition_draft_comment,
                    advertising_structures_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    advertising_structures_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    advertising_structures_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    advertising_structures_has_sample_document_comment AS has_sample_document_comment,
                    advertising_structures_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'capital_construction':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    capital_construction_id_rgmu AS id_rgmu,
                    capital_construction_statement_amount AS statement_amount,
                    capital_construction_link AS link,
                    capital_construction_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    capital_construction_has_btn_get_service_comment AS has_btn_get_service_comment,
                    capital_construction_has_reglament_comment AS has_reglament_comment,
                    capital_construction_has_estimation_quality_comment AS has_estimation_quality_comment,
                    capital_construction_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    capital_construction_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    capital_construction_has_edition_draft_comment AS has_edition_draft_comment,
                    capital_construction_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    capital_construction_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    capital_construction_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    capital_construction_has_sample_document_comment AS has_sample_document_comment,
                    capital_construction_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'preschool_education':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    preschool_education_id_rgmu AS id_rgmu,
                    preschool_education_statement_amount AS statement_amount,
                    preschool_education_link AS link,
                    preschool_education_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    preschool_education_has_btn_get_service_comment AS has_btn_get_service_comment,
                    preschool_education_has_reglament_comment AS has_reglament_comment,
                    preschool_education_has_estimation_quality_comment AS has_estimation_quality_comment,
                    preschool_education_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    preschool_education_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    preschool_education_has_edition_draft_comment AS has_edition_draft_comment,
                    preschool_education_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    preschool_education_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    preschool_education_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    preschool_education_has_sample_document_comment AS has_sample_document_comment,
                    preschool_education_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'school_education':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    school_education_id_rgmu AS id_rgmu,
                    school_education_statement_amount AS statement_amount,
                    school_education_link AS link,
                    school_education_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    school_education_has_btn_get_service_comment AS has_btn_get_service_comment,
                    school_education_has_reglament_comment AS has_reglament_comment,
                    school_education_has_estimation_quality_comment AS has_estimation_quality_comment,
                    school_education_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    school_education_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    school_education_has_edition_draft_comment AS has_edition_draft_comment,
                    school_education_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    school_education_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    school_education_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    school_education_has_sample_document_comment AS has_sample_document_comment,
                    school_education_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'needing_premises':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    needing_premises_id_rgmu AS id_rgmu,
                    needing_premises_statement_amount AS statement_amount,
                    needing_premises_link AS link,
                    needing_premises_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    needing_premises_has_btn_get_service_comment AS has_btn_get_service_comment,
                    needing_premises_has_reglament_comment AS has_reglament_comment,
                    needing_premises_has_estimation_quality_comment AS has_estimation_quality_comment,
                    needing_premises_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    needing_premises_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    needing_premises_has_edition_draft_comment AS has_edition_draft_comment,
                    needing_premises_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    needing_premises_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    needing_premises_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    needing_premises_has_sample_document_comment AS has_sample_document_comment,
                    needing_premises_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'town_planning':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    town_planning_id_rgmu AS id_rgmu,
                    town_planning_statement_amount AS statement_amount,
                    town_planning_link AS link,
                    town_planning_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    town_planning_has_btn_get_service_comment AS has_btn_get_service_comment,
                    town_planning_has_reglament_comment AS has_reglament_comment,
                    town_planning_has_estimation_quality_comment AS has_estimation_quality_comment,
                    town_planning_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    town_planning_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    town_planning_has_edition_draft_comment AS has_edition_draft_comment,
                    town_planning_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    town_planning_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    town_planning_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    town_planning_has_sample_document_comment AS has_sample_document_comment,
                    town_planning_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'archive_reference':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    archive_reference_id_rgmu AS id_rgmu,
                    archive_reference_statement_amount AS statement_amount,
                    archive_reference_link AS link,
                    archive_reference_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    archive_reference_has_btn_get_service_comment AS has_btn_get_service_comment,
                    archive_reference_has_reglament_comment AS has_reglament_comment,
                    archive_reference_has_estimation_quality_comment AS has_estimation_quality_comment,
                    archive_reference_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    archive_reference_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    archive_reference_has_edition_draft_comment AS has_edition_draft_comment,
                    archive_reference_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    archive_reference_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    archive_reference_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    archive_reference_has_sample_document_comment AS has_sample_document_comment,
                    archive_reference_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'land_schemes':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    land_schemes_id_rgmu AS id_rgmu,
                    land_schemes_statement_amount AS statement_amount,
                    land_schemes_link AS link,
                    land_schemes_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    land_schemes_has_btn_get_service_comment AS has_btn_get_service_comment,
                    land_schemes_has_reglament_comment AS has_reglament_comment,
                    land_schemes_has_estimation_quality_comment AS has_estimation_quality_comment,
                    land_schemes_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    land_schemes_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    land_schemes_has_edition_draft_comment AS has_edition_draft_comment,
                    land_schemes_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    land_schemes_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    land_schemes_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    land_schemes_has_sample_document_comment AS has_sample_document_comment,
                    land_schemes_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'land_sale':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    land_sale_id_rgmu AS id_rgmu,
                    land_sale_statement_amount AS statement_amount,
                    land_sale_link AS link,
                    land_sale_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    land_sale_has_btn_get_service_comment AS has_btn_get_service_comment,
                    land_sale_has_reglament_comment AS has_reglament_comment,
                    land_sale_has_estimation_quality_comment AS has_estimation_quality_comment,
                    land_sale_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    land_sale_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    land_sale_has_edition_draft_comment AS has_edition_draft_comment,
                    land_sale_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    land_sale_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    land_sale_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    land_sale_has_sample_document_comment AS has_sample_document_comment,
                    land_sale_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'land_lease':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    land_lease_id_rgmu AS id_rgmu,
                    land_lease_statement_amount AS statement_amount,
                    land_lease_link AS link,
                    land_lease_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    land_lease_has_btn_get_service_comment AS has_btn_get_service_comment,
                    land_lease_has_reglament_comment AS has_reglament_comment,
                    land_lease_has_estimation_quality_comment AS has_estimation_quality_comment,
                    land_lease_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    land_lease_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    land_lease_has_edition_draft_comment AS has_edition_draft_comment,
                    land_lease_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    land_lease_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    land_lease_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    land_lease_has_sample_document_comment AS has_sample_document_comment,
                    land_lease_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'ownership_right':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    ownership_right_id_rgmu AS id_rgmu,
                    ownership_right_statement_amount AS statement_amount,
                    ownership_right_link AS link,
                    ownership_right_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    ownership_right_has_btn_get_service_comment AS has_btn_get_service_comment,
                    ownership_right_has_reglament_comment AS has_reglament_comment,
                    ownership_right_has_estimation_quality_comment AS has_estimation_quality_comment,
                    ownership_right_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    ownership_right_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    ownership_right_has_edition_draft_comment AS has_edition_draft_comment,
                    ownership_right_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    ownership_right_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    ownership_right_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    ownership_right_has_sample_document_comment AS has_sample_document_comment,
                    ownership_right_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'municipal_property_lease':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    municipal_property_lease_id_rgmu AS id_rgmu,
                    municipal_property_lease_statement_amount AS statement_amount,
                    municipal_property_lease_link AS link,
                    municipal_property_lease_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    municipal_property_lease_has_btn_get_service_comment AS has_btn_get_service_comment,
                    municipal_property_lease_has_reglament_comment AS has_reglament_comment,
                    municipal_property_lease_has_estimation_quality_comment AS has_estimation_quality_comment,
                    municipal_property_lease_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    municipal_property_lease_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    municipal_property_lease_has_edition_draft_comment AS has_edition_draft_comment,
                    municipal_property_lease_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    municipal_property_lease_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    municipal_property_lease_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    municipal_property_lease_has_sample_document_comment AS has_sample_document_comment,
                    municipal_property_lease_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    elif service_name == 'free_land_provision':
        for name in regions_names:
            try:
                obj = RegionModel.objects.raw(
                    '''SELECT id, region_name,
                    free_land_provision_id_rgmu AS id_rgmu,
                    free_land_provision_statement_amount AS statement_amount,
                    free_land_provision_link AS link,
                    free_land_provision_has_advanced_appointment_comment AS has_advanced_appointment_comment,
                    free_land_provision_has_btn_get_service_comment AS has_btn_get_service_comment,
                    free_land_provision_has_reglament_comment AS has_reglament_comment,
                    free_land_provision_has_estimation_quality_comment AS has_estimation_quality_comment,
                    free_land_provision_connected_to_fgis_do_comment AS connected_to_fgis_do_comment,
                    free_land_provision_has_electronic_form_printing_comment AS has_electronic_form_printing_comment,
                    free_land_provision_has_edition_draft_comment AS has_edition_draft_comment,
                    free_land_provision_has_term_of_consideration_comment AS has_term_of_consideration_comment,
                    free_land_provision_has_notif_consider_result_comment AS has_notif_consider_result_comment,
                    free_land_provision_has_causes_of_failure_comment AS has_causes_of_failure_comment,
                    free_land_provision_has_sample_document_comment AS has_sample_document_comment,
                    free_land_provision_has_document_template_comment AS has_document_template_comment
                    FROM app_regionModel
                    WHERE region_name =\'''' + str(name) + '''\' AND month=\'''' + str(
                        month) + '''\' AND year=\'''' + str(
                        year) + '''\' ORDER BY year DESC, month DESC, day DESC, time DESC LIMIT 1;''')[0]
                objects.append(obj)
            except IndexError:
                obj = {'region_name': name, 'id_rgmu': "", 'statement_amount': "", 'link': "",
                       'has_advanced_appointment_comment': "", 'has_btn_get_service_comment': "",
                       'has_reglament_comment': "", 'has_estimation_quality_comment': "",
                       'connected_to_fgis_do_comment': "", 'has_electronic_form_printing_comment': "",
                       'has_edition_draft_comment': "", 'has_term_of_consideration_comment': "",
                       'has_notif_consider_result_comment': "", 'has_causes_of_failure_comment': "",
                       'has_sample_document_comment': "", 'has_document_template_comment': ""}
                objects.append(dotDict(obj))

    else:
        return HttpResponseNotFound('')

    return render(request, 'app/result_form.html', {'objects': objects,
                                                        'year': str(year),
                                                        'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                        'num_month': month,
                                                        'full_service_name': full_service_names[short_service_names.index(service_name)],
                                                        'service_name': service_name,
                                                        'zipped': zip(regions_names, short_regions_names),
                                                        'username': request.user.username,
                                                        'years': [i for i in range(2016, int(datetime.now().year)+1)],
                                                        'zipped_service_names': zip(full_service_names, short_service_names)})


@login_required(login_url='/login/',
                redirect_field_name='/result_form/' + datetime.today().strftime('%Y/%m/')+'ufa/')
def get_region_form(request, year, month, short_region_name):
    full_region_name = regions_names[short_regions_names.index(short_region_name)]
    try:
        object = get_object_or_404(RegionModel, region_name=full_region_name, month=month, year=year)

    except Http404:
        object = {'short_region_name': short_region_name,
                  'residential_premises_id_rgmu': "",
                  'residential_premises_statement_amount': "",
                  'residential_premises_link': "",
                  'residential_premises_has_advanced_appointment_comment': "",
                  'residential_premises_has_btn_get_service_comment': "",
                  'residential_premises_has_reglament_comment': "",
                  'residential_premises_has_estimation_quality_comment': "",
                  'residential_premises_connected_to_fgis_do_comment': "",
                  'residential_premises_has_electronic_form_printing_comment': "",
                  'residential_premises_has_edition_draft_comment': "",
                  'residential_premises_has_term_of_consideration_comment': "",
                  'residential_premises_has_notif_consider_result_comment': "",
                  'residential_premises_has_causes_of_failure_comment': "",
                  'residential_premises_has_sample_document_comment': "",
                  'residential_premises_has_document_template_comment': "",
                  'housing_transfer_id_rgmu': "",
                  'housing_transfer_statement_amount': "",
                  'housing_transfer_link': "",
                  'housing_transfer_has_advanced_appointment_comment': "",
                  'housing_transfer_has_btn_get_service_comment': "",
                  'housing_transfer_has_reglament_comment': "",
                  'housing_transfer_has_estimation_quality_comment': "",
                  'housing_transfer_connected_to_fgis_do_comment': "",
                  'housing_transfer_has_electronic_form_printing_comment': "",
                  'housing_transfer_has_edition_draft_comment': "",
                  'housing_transfer_has_term_of_consideration_comment': "",
                  'housing_transfer_has_notif_consider_result_comment': "",
                  'housing_transfer_has_causes_of_failure_comment': "",
                  'housing_transfer_has_sample_document_comment': "",
                  'housing_transfer_has_document_template_comment': "",
                  'advertising_structures_id_rgmu': "",
                  'advertising_structures_statement_amount': "",
                  'advertising_structures_link': "",
                  'advertising_structures_has_advanced_appointment_comment': "",
                  'advertising_structures_has_btn_get_service_comment': "",
                  'advertising_structures_has_reglament_comment': "",
                  'advertising_structures_has_estimation_quality_comment': "",
                  'advertising_structures_connected_to_fgis_do_comment': "",
                  'advertising_structures_has_electronic_form_printing_comment': "",
                  'advertising_structures_has_edition_draft_comment': "",
                  'advertising_structures_has_term_of_consideration_comment': "",
                  'advertising_structures_has_notif_consider_result_comment': "",
                  'advertising_structures_has_causes_of_failure_comment': "",
                  'advertising_structures_has_sample_document_comment': "",
                  'advertising_structures_has_document_template_comment': "",
                  'capital_construction_id_rgmu': "",
                  'capital_construction_statement_amount': "",
                  'capital_construction_link': "",
                  'capital_construction_has_advanced_appointment_comment': "",
                  'capital_construction_has_btn_get_service_comment': "",
                  'capital_construction_has_reglament_comment': "",
                  'capital_construction_has_estimation_quality_comment': "",
                  'capital_construction_connected_to_fgis_do_comment': "",
                  'capital_construction_has_electronic_form_printing_comment': "",
                  'capital_construction_has_edition_draft_comment': "",
                  'capital_construction_has_term_of_consideration_comment': "",
                  'capital_construction_has_notif_consider_result_comment': "",
                  'capital_construction_has_causes_of_failure_comment': "",
                  'capital_construction_has_sample_document_comment': "",
                  'capital_construction_has_document_template_comment': "",
                  'preschool_education_id_rgmu': "",
                  'preschool_education_statement_amount': "",
                  'preschool_education_link': "",
                  'preschool_education_has_advanced_appointment_comment': "",
                  'preschool_education_has_btn_get_service_comment': "",
                  'preschool_education_has_reglament_comment': "",
                  'preschool_education_has_estimation_quality_comment': "",
                  'preschool_education_connected_to_fgis_do_comment': "",
                  'preschool_education_has_electronic_form_printing_comment': "",
                  'preschool_education_has_edition_draft_comment': "",
                  'preschool_education_has_term_of_consideration_comment': "",
                  'preschool_education_has_notif_consider_result_comment': "",
                  'preschool_education_has_causes_of_failure_comment': "",
                  'preschool_education_has_sample_document_comment': "",
                  'preschool_education_has_document_template_comment': "",
                  'school_education_id_rgmu': "",
                  'school_education_statement_amount': "",
                  'school_education_link': "",
                  'school_education_has_advanced_appointment_comment': "",
                  'school_education_has_btn_get_service_comment': "",
                  'school_education_has_reglament_comment': "",
                  'school_education_has_estimation_quality_comment': "",
                  'school_education_connected_to_fgis_do_comment': "",
                  'school_education_has_electronic_form_printing_comment': "",
                  'school_education_has_edition_draft_comment': "",
                  'school_education_has_term_of_consideration_comment': "",
                  'school_education_has_notif_consider_result_comment': "",
                  'school_education_has_causes_of_failure_comment': "",
                  'school_education_has_sample_document_comment': "",
                  'school_education_has_document_template_comment': "",
                  'needing_premises_id_rgmu': "",
                  'needing_premises_statement_amount': "",
                  'needing_premises_link': "",
                  'needing_premises_has_advanced_appointment_comment': "",
                  'needing_premises_has_btn_get_service_comment': "",
                  'needing_premises_has_reglament_comment': "",
                  'needing_premises_has_estimation_quality_comment': "",
                  'needing_premises_connected_to_fgis_do_comment': "",
                  'needing_premises_has_electronic_form_printing_comment': "",
                  'needing_premises_has_edition_draft_comment': "",
                  'needing_premises_has_term_of_consideration_comment': "",
                  'needing_premises_has_notif_consider_result_comment': "",
                  'needing_premises_has_causes_of_failure_comment': "",
                  'needing_premises_has_sample_document_comment': "",
                  'needing_premises_has_document_template_comment': "",
                  'town_planning_id_rgmu': "",
                  'town_planning_statement_amount': "",
                  'town_planning_link': "",
                  'town_planning_has_advanced_appointment_comment': "",
                  'town_planning_has_btn_get_service_comment': "",
                  'town_planning_has_reglament_comment': "",
                  'town_planning_has_estimation_quality_comment': "",
                  'town_planning_connected_to_fgis_do_comment': "",
                  'town_planning_has_electronic_form_printing_comment': "",
                  'town_planning_has_edition_draft_comment': "",
                  'town_planning_has_term_of_consideration_comment': "",
                  'town_planning_has_notif_consider_result_comment': "",
                  'town_planning_has_causes_of_failure_comment': "",
                  'town_planning_has_sample_document_comment': "",
                  'town_planning_has_document_template_comment': "",
                  'archive_reference_id_rgmu': "",
                  'archive_reference_statement_amount': "",
                  'archive_reference_link': "",
                  'archive_reference_has_advanced_appointment_comment': "",
                  'archive_reference_has_btn_get_service_comment': "",
                  'archive_reference_has_reglament_comment': "",
                  'archive_reference_has_estimation_quality_comment': "",
                  'archive_reference_connected_to_fgis_do_comment': "",
                  'archive_reference_has_electronic_form_printing_comment': "",
                  'archive_reference_has_edition_draft_comment': "",
                  'archive_reference_has_term_of_consideration_comment': "",
                  'archive_reference_has_notif_consider_result_comment': "",
                  'archive_reference_has_causes_of_failure_comment': "",
                  'archive_reference_has_sample_document_comment': "",
                  'archive_reference_has_document_template_comment': "",
                  'land_schemes_id_rgmu': "",
                  'land_schemes_statement_amount': "",
                  'land_schemes_link': "",
                  'land_schemes_has_advanced_appointment_comment': "",
                  'land_schemes_has_btn_get_service_comment': "",
                  'land_schemes_has_reglament_comment': "",
                  'land_schemes_has_estimation_quality_comment': "",
                  'land_schemes_connected_to_fgis_do_comment': "",
                  'land_schemes_has_electronic_form_printing_comment': "",
                  'land_schemes_has_edition_draft_comment': "",
                  'land_schemes_has_term_of_consideration_comment': "",
                  'land_schemes_has_notif_consider_result_comment': "",
                  'land_schemes_has_causes_of_failure_comment': "",
                  'land_schemes_has_sample_document_comment': "",
                  'land_schemes_has_document_template_comment': "",
                  'land_sale_id_rgmu': "",
                  'land_sale_statement_amount': "",
                  'land_sale_link': "",
                  'land_sale_has_advanced_appointment_comment': "",
                  'land_sale_has_btn_get_service_comment': "",
                  'land_sale_has_reglament_comment': "",
                  'land_sale_has_estimation_quality_comment': "",
                  'land_sale_connected_to_fgis_do_comment': "",
                  'land_sale_has_electronic_form_printing_comment': "",
                  'land_sale_has_edition_draft_comment': "",
                  'land_sale_has_term_of_consideration_comment': "",
                  'land_sale_has_notif_consider_result_comment': "",
                  'land_sale_has_causes_of_failure_comment': "",
                  'land_sale_has_sample_document_comment': "",
                  'land_sale_has_document_template_comment': "",
                  'land_lease_id_rgmu': "",
                  'land_lease_statement_amount': "",
                  'land_lease_link': "",
                  'land_lease_has_advanced_appointment_comment': "",
                  'land_lease_has_btn_get_service_comment': "",
                  'land_lease_has_reglament_comment': "",
                  'land_lease_has_estimation_quality_comment': "",
                  'land_lease_connected_to_fgis_do_comment': "",
                  'land_lease_has_electronic_form_printing_comment': "",
                  'land_lease_has_edition_draft_comment': "",
                  'land_lease_has_term_of_consideration_comment': "",
                  'land_lease_has_notif_consider_result_comment': "",
                  'land_lease_has_causes_of_failure_comment': "",
                  'land_lease_has_sample_document_comment': "",
                  'land_lease_has_document_template_comment': "",
                  'ownership_right_id_rgmu': "",
                  'ownership_right_statement_amount': "",
                  'ownership_right_link': "",
                  'ownership_right_has_advanced_appointment_comment': "",
                  'ownership_right_has_btn_get_service_comment': "",
                  'ownership_right_has_reglament_comment': "",
                  'ownership_right_has_estimation_quality_comment': "",
                  'ownership_right_connected_to_fgis_do_comment': "",
                  'ownership_right_has_electronic_form_printing_comment': "",
                  'ownership_right_has_edition_draft_comment': "",
                  'ownership_right_has_term_of_consideration_comment': "",
                  'ownership_right_has_notif_consider_result_comment': "",
                  'ownership_right_has_causes_of_failure_comment': "",
                  'ownership_right_has_sample_document_comment': "",
                  'ownership_right_has_document_template_comment': "",
                  'municipal_property_lease_id_rgmu': "",
                  'municipal_property_lease_statement_amount': "",
                  'municipal_property_lease_link': "",
                  'municipal_property_lease_has_advanced_appointment_comment': "",
                  'municipal_property_lease_has_btn_get_service_comment': "",
                  'municipal_property_lease_has_reglament_comment': "",
                  'municipal_property_lease_has_estimation_quality_comment': "",
                  'municipal_property_lease_connected_to_fgis_do_comment': "",
                  'municipal_property_lease_has_electronic_form_printing_comment': "",
                  'municipal_property_lease_has_edition_draft_comment': "",
                  'municipal_property_lease_has_term_of_consideration_comment': "",
                  'municipal_property_lease_has_notif_consider_result_comment': "",
                  'municipal_property_lease_has_causes_of_failure_comment': "",
                  'municipal_property_lease_has_sample_document_comment': "",
                  'municipal_property_lease_has_document_template_comment': "",
                  'free_land_provision_id_rgmu': "",
                  'free_land_provision_statement_amount': "",
                  'free_land_provision_link': "",
                  'free_land_provision_has_advanced_appointment_comment': "",
                  'free_land_provision_has_btn_get_service_comment': "",
                  'free_land_provision_has_reglament_comment': "",
                  'free_land_provision_has_estimation_quality_comment': "",
                  'free_land_provision_connected_to_fgis_do_comment': "",
                  'free_land_provision_has_electronic_form_printing_comment': "",
                  'free_land_provision_has_edition_draft_comment': "",
                  'free_land_provision_has_term_of_consideration_comment': "",
                  'free_land_provision_has_notif_consider_result_comment': "",
                  'free_land_provision_has_causes_of_failure_comment': "",
                  'free_land_provision_has_sample_document_comment': "",
                  'free_land_provision_has_document_template_comment': "",
                  }
    return render(request, 'app/region_form.html', {'object': object,
                                                        'year': str(year),
                                                        'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                        'num_month': month,
                                                        'zipped': zip(regions_names, short_regions_names),
                                                        'full_region_name': full_region_name,
                                                        'short_region_name': short_region_name,
                                                        'username': request.user.username,
                                                        'years': [i for i in range(2016, int(datetime.now().year) + 1)],
                                                        'zipped_service_names': zip(full_service_names, short_service_names)
                                                        })


@login_required(login_url='/login/',
                redirect_field_name='/result_form/with_troubles/' + datetime.today().strftime('%Y/%m/'))
def get_result_form_with_troubles(request, year, month):
    objects = get_with_troubles(month,year)
    return render(request, 'app/with_troubles.html', {'objects': dotDict(objects),
                                                        'year': str(year),
                                                        'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                        'num_month': month,
                                                        'zipped': zip(regions_names, short_regions_names),
                                                        'username': request.user.username,
                                                        'years': [i for i in range(2016, int(datetime.now().year) + 1)],
                                                        'zipped_service_names': zip(full_service_names,
                                                                                  short_service_names)})


@login_required(login_url='/login/',
                redirect_field_name='/result_form/not_sent/' + datetime.today().strftime('%Y/%m/'))
def get_result_form_not_sent(request, year, month):
    region_full_names = get_not_sent(month,year)
    indices = [regions_names.index(full_name) for full_name in region_full_names]
    return render(request, 'app/not_sent.html', {
                                                    'zipped_names': zip(region_full_names, [short_regions_names[i] for i in indices]),
                                                    'year': str(year),
                                                     'month': MONTHS[MONTH_NUMBERS.index(str(month))],
                                                     'num_month': month,
                                                     'zipped': zip(regions_names, short_regions_names),
                                                     'username': request.user.username,
                                                     'years': [i for i in range(2016, int(datetime.now().year) + 1)],
                                                    'zipped_service_names': zip(full_service_names, short_service_names)
                                                     })


def login_view(request):
    form = LoginForm(request.POST or None)
    print("form created", request.POST,form.is_valid())
    if request.POST and form.is_valid():
        user = form.login(request)
        if user:
            login(request, user)
            return HttpResponseRedirect(reverse('result_form', kwargs={
                'service_name': 'residential_premises',
                'year': datetime.today().strftime('%Y'),
                'month': datetime.today().strftime('%m')
            }))
    return render(request, 'app/registration/login.html', {'form': form})


def logout_view(request):
   logout(request)
   return HttpResponseRedirect('/login/')

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def export_all(request,year,month):
    if request.method == 'GET':
        wb = load_workbook('./app/static/app/xlsx_templates/all_data.xlsx')

        ws1 = wb["Лист 1"]
        ws2 = wb["Лист 2"]
        ws3 = wb["Лист 3"]
        ws4 = wb["Лист 4"]
        ws5 = wb["Лист 5"]
        ws6 = wb["Лист 6"]
        ws7 = wb["Лист 7"]
        ws8 = wb["Лист 8"]
        ws9 = wb["Лист 9"]
        ws10 = wb["Лист 10"]
        ws11 = wb["Лист 11"]
        ws12 = wb["Лист 12"]
        ws13 = wb["Лист 13"]
        ws14 = wb["Лист 14"]
        ws15 = wb["Лист 15"]
        ws1[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws1['A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[0], month, year)
        ws2['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws2[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[1], month, year)
        ws3[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws3[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[2], month, year)
        ws4[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws4[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[3], month, year)
        ws5[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws5[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[4], month, year)
        ws6[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws6[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[5], month, year)
        ws7[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws7[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[6], month, year)
        ws8[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws8[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[7], month, year)
        ws9[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в (количество указывается нарастающим итогом)'.format(
            month, year)
        ws9[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[8], month, year)
        ws10[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws10[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[9], month, year)
        ws11[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws11[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[10], month, year)
        ws12[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws12[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[11], month, year)
        ws13[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws13[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[12], month, year)
        ws14[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws14[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[13], month, year)
        ws15[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws15[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме во всех районах Республики Башкортостан за {}.{}г.'.format(
            full_service_names[14], month, year)

        objects = []
        for name in regions_names:
            try:
                object = get_object_or_404(RegionModel,region_name=name, year=year, month=month)
                objects.append(object)
            except Http404:
                object = {'region_name': name,
                          'residential_premises_id_rgmu': "",
                          'residential_premises_statement_amount': "",
                          'residential_premises_link': "",
                          'residential_premises_has_advanced_appointment_comment': "",
                          'residential_premises_has_btn_get_service_comment': "",
                          'residential_premises_has_reglament_comment': "",
                          'residential_premises_has_estimation_quality_comment': "",
                          'residential_premises_connected_to_fgis_do_comment': "",
                          'residential_premises_has_electronic_form_printing_comment': "",
                          'residential_premises_has_edition_draft_comment': "",
                          'residential_premises_has_term_of_consideration_comment': "",
                          'residential_premises_has_notif_consider_result_comment': "",
                          'residential_premises_has_causes_of_failure_comment': "",
                          'residential_premises_has_sample_document_comment': "",
                          'residential_premises_has_document_template_comment': "",
                          'housing_transfer_id_rgmu': "",
                          'housing_transfer_statement_amount': "",
                          'housing_transfer_link': "",
                          'housing_transfer_has_advanced_appointment_comment': "",
                          'housing_transfer_has_btn_get_service_comment': "",
                          'housing_transfer_has_reglament_comment': "",
                          'housing_transfer_has_estimation_quality_comment': "",
                          'housing_transfer_connected_to_fgis_do_comment': "",
                          'housing_transfer_has_electronic_form_printing_comment': "",
                          'housing_transfer_has_edition_draft_comment': "",
                          'housing_transfer_has_term_of_consideration_comment': "",
                          'housing_transfer_has_notif_consider_result_comment': "",
                          'housing_transfer_has_causes_of_failure_comment': "",
                          'housing_transfer_has_sample_document_comment': "",
                          'housing_transfer_has_document_template_comment': "",
                          'advertising_structures_id_rgmu': "",
                          'advertising_structures_statement_amount': "",
                          'advertising_structures_link': "",
                          'advertising_structures_has_advanced_appointment_comment': "",
                          'advertising_structures_has_btn_get_service_comment': "",
                          'advertising_structures_has_reglament_comment': "",
                          'advertising_structures_has_estimation_quality_comment': "",
                          'advertising_structures_connected_to_fgis_do_comment': "",
                          'advertising_structures_has_electronic_form_printing_comment': "",
                          'advertising_structures_has_edition_draft_comment': "",
                          'advertising_structures_has_term_of_consideration_comment': "",
                          'advertising_structures_has_notif_consider_result_comment': "",
                          'advertising_structures_has_causes_of_failure_comment': "",
                          'advertising_structures_has_sample_document_comment': "",
                          'advertising_structures_has_document_template_comment': "",
                          'capital_construction_id_rgmu': "",
                          'capital_construction_statement_amount': "",
                          'capital_construction_link': "",
                          'capital_construction_has_advanced_appointment_comment': "",
                          'capital_construction_has_btn_get_service_comment': "",
                          'capital_construction_has_reglament_comment': "",
                          'capital_construction_has_estimation_quality_comment': "",
                          'capital_construction_connected_to_fgis_do_comment': "",
                          'capital_construction_has_electronic_form_printing_comment': "",
                          'capital_construction_has_edition_draft_comment': "",
                          'capital_construction_has_term_of_consideration_comment': "",
                          'capital_construction_has_notif_consider_result_comment': "",
                          'capital_construction_has_causes_of_failure_comment': "",
                          'capital_construction_has_sample_document_comment': "",
                          'capital_construction_has_document_template_comment': "",
                          'preschool_education_id_rgmu': "",
                          'preschool_education_statement_amount': "",
                          'preschool_education_link': "",
                          'preschool_education_has_advanced_appointment_comment': "",
                          'preschool_education_has_btn_get_service_comment': "",
                          'preschool_education_has_reglament_comment': "",
                          'preschool_education_has_estimation_quality_comment': "",
                          'preschool_education_connected_to_fgis_do_comment': "",
                          'preschool_education_has_electronic_form_printing_comment': "",
                          'preschool_education_has_edition_draft_comment': "",
                          'preschool_education_has_term_of_consideration_comment': "",
                          'preschool_education_has_notif_consider_result_comment': "",
                          'preschool_education_has_causes_of_failure_comment': "",
                          'preschool_education_has_sample_document_comment': "",
                          'preschool_education_has_document_template_comment': "",
                          'school_education_id_rgmu': "",
                          'school_education_statement_amount': "",
                          'school_education_link': "",
                          'school_education_has_advanced_appointment_comment': "",
                          'school_education_has_btn_get_service_comment': "",
                          'school_education_has_reglament_comment': "",
                          'school_education_has_estimation_quality_comment': "",
                          'school_education_connected_to_fgis_do_comment': "",
                          'school_education_has_electronic_form_printing_comment': "",
                          'school_education_has_edition_draft_comment': "",
                          'school_education_has_term_of_consideration_comment': "",
                          'school_education_has_notif_consider_result_comment': "",
                          'school_education_has_causes_of_failure_comment': "",
                          'school_education_has_sample_document_comment': "",
                          'school_education_has_document_template_comment': "",
                          'needing_premises_id_rgmu': "",
                          'needing_premises_statement_amount': "",
                          'needing_premises_link': "",
                          'needing_premises_has_advanced_appointment_comment': "",
                          'needing_premises_has_btn_get_service_comment': "",
                          'needing_premises_has_reglament_comment': "",
                          'needing_premises_has_estimation_quality_comment': "",
                          'needing_premises_connected_to_fgis_do_comment': "",
                          'needing_premises_has_electronic_form_printing_comment': "",
                          'needing_premises_has_edition_draft_comment': "",
                          'needing_premises_has_term_of_consideration_comment': "",
                          'needing_premises_has_notif_consider_result_comment': "",
                          'needing_premises_has_causes_of_failure_comment': "",
                          'needing_premises_has_sample_document_comment': "",
                          'needing_premises_has_document_template_comment': "",
                          'town_planning_id_rgmu': "",
                          'town_planning_statement_amount': "",
                          'town_planning_link': "",
                          'town_planning_has_advanced_appointment_comment': "",
                          'town_planning_has_btn_get_service_comment': "",
                          'town_planning_has_reglament_comment': "",
                          'town_planning_has_estimation_quality_comment': "",
                          'town_planning_connected_to_fgis_do_comment': "",
                          'town_planning_has_electronic_form_printing_comment': "",
                          'town_planning_has_edition_draft_comment': "",
                          'town_planning_has_term_of_consideration_comment': "",
                          'town_planning_has_notif_consider_result_comment': "",
                          'town_planning_has_causes_of_failure_comment': "",
                          'town_planning_has_sample_document_comment': "",
                          'town_planning_has_document_template_comment': "",
                          'archive_reference_id_rgmu': "",
                          'archive_reference_statement_amount': "",
                          'archive_reference_link': "",
                          'archive_reference_has_advanced_appointment_comment': "",
                          'archive_reference_has_btn_get_service_comment': "",
                          'archive_reference_has_reglament_comment': "",
                          'archive_reference_has_estimation_quality_comment': "",
                          'archive_reference_connected_to_fgis_do_comment': "",
                          'archive_reference_has_electronic_form_printing_comment': "",
                          'archive_reference_has_edition_draft_comment': "",
                          'archive_reference_has_term_of_consideration_comment': "",
                          'archive_reference_has_notif_consider_result_comment': "",
                          'archive_reference_has_causes_of_failure_comment': "",
                          'archive_reference_has_sample_document_comment': "",
                          'archive_reference_has_document_template_comment': "",
                          'land_schemes_id_rgmu': "",
                          'land_schemes_statement_amount': "",
                          'land_schemes_link': "",
                          'land_schemes_has_advanced_appointment_comment': "",
                          'land_schemes_has_btn_get_service_comment': "",
                          'land_schemes_has_reglament_comment': "",
                          'land_schemes_has_estimation_quality_comment': "",
                          'land_schemes_connected_to_fgis_do_comment': "",
                          'land_schemes_has_electronic_form_printing_comment': "",
                          'land_schemes_has_edition_draft_comment': "",
                          'land_schemes_has_term_of_consideration_comment': "",
                          'land_schemes_has_notif_consider_result_comment': "",
                          'land_schemes_has_causes_of_failure_comment': "",
                          'land_schemes_has_sample_document_comment': "",
                          'land_schemes_has_document_template_comment': "",
                          'land_sale_id_rgmu': "",
                          'land_sale_statement_amount': "",
                          'land_sale_link': "",
                          'land_sale_has_advanced_appointment_comment': "",
                          'land_sale_has_btn_get_service_comment': "",
                          'land_sale_has_reglament_comment': "",
                          'land_sale_has_estimation_quality_comment': "",
                          'land_sale_connected_to_fgis_do_comment': "",
                          'land_sale_has_electronic_form_printing_comment': "",
                          'land_sale_has_edition_draft_comment': "",
                          'land_sale_has_term_of_consideration_comment': "",
                          'land_sale_has_notif_consider_result_comment': "",
                          'land_sale_has_causes_of_failure_comment': "",
                          'land_sale_has_sample_document_comment': "",
                          'land_sale_has_document_template_comment': "",
                          'land_lease_id_rgmu': "",
                          'land_lease_statement_amount': "",
                          'land_lease_link': "",
                          'land_lease_has_advanced_appointment_comment': "",
                          'land_lease_has_btn_get_service_comment': "",
                          'land_lease_has_reglament_comment': "",
                          'land_lease_has_estimation_quality_comment': "",
                          'land_lease_connected_to_fgis_do_comment': "",
                          'land_lease_has_electronic_form_printing_comment': "",
                          'land_lease_has_edition_draft_comment': "",
                          'land_lease_has_term_of_consideration_comment': "",
                          'land_lease_has_notif_consider_result_comment': "",
                          'land_lease_has_causes_of_failure_comment': "",
                          'land_lease_has_sample_document_comment': "",
                          'land_lease_has_document_template_comment': "",
                          'ownership_right_id_rgmu': "",
                          'ownership_right_statement_amount': "",
                          'ownership_right_link': "",
                          'ownership_right_has_advanced_appointment_comment': "",
                          'ownership_right_has_btn_get_service_comment': "",
                          'ownership_right_has_reglament_comment': "",
                          'ownership_right_has_estimation_quality_comment': "",
                          'ownership_right_connected_to_fgis_do_comment': "",
                          'ownership_right_has_electronic_form_printing_comment': "",
                          'ownership_right_has_edition_draft_comment': "",
                          'ownership_right_has_term_of_consideration_comment': "",
                          'ownership_right_has_notif_consider_result_comment': "",
                          'ownership_right_has_causes_of_failure_comment': "",
                          'ownership_right_has_sample_document_comment': "",
                          'ownership_right_has_document_template_comment': "",
                          'municipal_property_lease_id_rgmu': "",
                          'municipal_property_lease_statement_amount': "",
                          'municipal_property_lease_link': "",
                          'municipal_property_lease_has_advanced_appointment_comment': "",
                          'municipal_property_lease_has_btn_get_service_comment': "",
                          'municipal_property_lease_has_reglament_comment': "",
                          'municipal_property_lease_has_estimation_quality_comment': "",
                          'municipal_property_lease_connected_to_fgis_do_comment': "",
                          'municipal_property_lease_has_electronic_form_printing_comment': "",
                          'municipal_property_lease_has_edition_draft_comment': "",
                          'municipal_property_lease_has_term_of_consideration_comment': "",
                          'municipal_property_lease_has_notif_consider_result_comment': "",
                          'municipal_property_lease_has_causes_of_failure_comment': "",
                          'municipal_property_lease_has_sample_document_comment': "",
                          'municipal_property_lease_has_document_template_comment': "",
                          'free_land_provision_id_rgmu': "",
                          'free_land_provision_statement_amount': "",
                          'free_land_provision_link': "",
                          'free_land_provision_has_advanced_appointment_comment': "",
                          'free_land_provision_has_btn_get_service_comment': "",
                          'free_land_provision_has_reglament_comment': "",
                          'free_land_provision_has_estimation_quality_comment': "",
                          'free_land_provision_connected_to_fgis_do_comment': "",
                          'free_land_provision_has_electronic_form_printing_comment': "",
                          'free_land_provision_has_edition_draft_comment': "",
                          'free_land_provision_has_term_of_consideration_comment': "",
                          'free_land_provision_has_notif_consider_result_comment': "",
                          'free_land_provision_has_causes_of_failure_comment': "",
                          'free_land_provision_has_sample_document_comment': "",
                          'free_land_provision_has_document_template_comment': "",
                          }
                objects.append(dotDict(object))
        for i in range(len(objects)):
                object = objects[i]
                ws1['C' + str(7+i)] = object.residential_premises_id_rgmu
                ws1['D' + str(7+i)] = object.residential_premises_statement_amount
                ws1['E' + str(7+i)] = object.residential_premises_link
                ws1['F' + str(7+i)] = object.residential_premises_has_advanced_appointment_comment
                ws1['G' + str(7+i)] = object.residential_premises_has_btn_get_service_comment
                ws1['H' + str(7+i)] = object.residential_premises_has_reglament_comment
                ws1['I' + str(7+i)] = object.residential_premises_has_estimation_quality_comment
                ws1['J' + str(7+i)] = object.residential_premises_connected_to_fgis_do_comment
                ws1['K' + str(7+i)] = object.residential_premises_has_electronic_form_printing_comment
                ws1['L' + str(7+i)] = object.residential_premises_has_edition_draft_comment
                ws1['M' + str(7+i)] = object.residential_premises_has_term_of_consideration_comment
                ws1['N' + str(7+i)] = object.residential_premises_has_notif_consider_result_comment
                ws1['O' + str(7+i)] = object.residential_premises_has_causes_of_failure_comment
                ws1['P' + str(7+i)] = object.residential_premises_has_sample_document_comment
                ws1['Q' + str(7+i)] = object.residential_premises_has_document_template_comment
                ws2['C' + str(7 + i)] = object.housing_transfer_id_rgmu
                ws2['D' + str(7 + i)] = object.housing_transfer_statement_amount
                ws2['E' + str(7 + i)] = object.housing_transfer_link
                ws2['F' + str(7 + i)] = object.housing_transfer_has_advanced_appointment_comment
                ws2['G' + str(7 + i)] = object.housing_transfer_has_btn_get_service_comment
                ws2['H' + str(7 + i)] = object.housing_transfer_has_reglament_comment
                ws2['I' + str(7 + i)] = object.housing_transfer_has_estimation_quality_comment
                ws2['J' + str(7 + i)] = object.housing_transfer_connected_to_fgis_do_comment
                ws2['K' + str(7 + i)] = object.housing_transfer_has_electronic_form_printing_comment
                ws2['L' + str(7 + i)] = object.housing_transfer_has_edition_draft_comment
                ws2['M' + str(7 + i)] = object.housing_transfer_has_term_of_consideration_comment
                ws2['N' + str(7 + i)] = object.housing_transfer_has_notif_consider_result_comment
                ws2['O' + str(7 + i)] = object.housing_transfer_has_causes_of_failure_comment
                ws2['P' + str(7 + i)] = object.housing_transfer_has_sample_document_comment
                ws2['Q' + str(7 + i)] = object.housing_transfer_has_document_template_comment
                ws3['C' + str(7 + i)] = object.advertising_structures_id_rgmu
                ws3['D' + str(7 + i)] = object.advertising_structures_statement_amount
                ws3['E' + str(7 + i)] = object.advertising_structures_link
                ws3['F' + str(7 + i)] = object.advertising_structures_has_advanced_appointment_comment
                ws3['G' + str(7 + i)] = object.advertising_structures_has_btn_get_service_comment
                ws3['H' + str(7 + i)] = object.advertising_structures_has_reglament_comment
                ws3['I' + str(7 + i)] = object.advertising_structures_has_estimation_quality_comment
                ws3['J' + str(7 + i)] = object.advertising_structures_connected_to_fgis_do_comment
                ws3['K' + str(7 + i)] = object.advertising_structures_has_electronic_form_printing_comment
                ws3['L' + str(7 + i)] = object.advertising_structures_has_edition_draft_comment
                ws3['M' + str(7 + i)] = object.advertising_structures_has_term_of_consideration_comment
                ws3['N' + str(7 + i)] = object.advertising_structures_has_notif_consider_result_comment
                ws3['O' + str(7 + i)] = object.advertising_structures_has_causes_of_failure_comment
                ws3['P' + str(7 + i)] = object.advertising_structures_has_sample_document_comment
                ws3['Q' + str(7 + i)] = object.advertising_structures_has_document_template_comment
                ws4['C' + str(7 + i)] = object.capital_construction_id_rgmu
                ws4['D' + str(7 + i)] = object.capital_construction_statement_amount
                ws4['E' + str(7 + i)] = object.capital_construction_link
                ws4['F' + str(7 + i)] = object.capital_construction_has_advanced_appointment_comment
                ws4['G' + str(7 + i)] = object.capital_construction_has_btn_get_service_comment
                ws4['H' + str(7 + i)] = object.capital_construction_has_reglament_comment
                ws4['I' + str(7 + i)] = object.capital_construction_has_estimation_quality_comment
                ws4['J' + str(7 + i)] = object.capital_construction_connected_to_fgis_do_comment
                ws4['K' + str(7 + i)] = object.capital_construction_has_electronic_form_printing_comment
                ws4['L' + str(7 + i)] = object.capital_construction_has_edition_draft_comment
                ws4['M' + str(7 + i)] = object.capital_construction_has_term_of_consideration_comment
                ws4['N' + str(7 + i)] = object.capital_construction_has_notif_consider_result_comment
                ws4['O' + str(7 + i)] = object.capital_construction_has_causes_of_failure_comment
                ws4['P' + str(7 + i)] = object.capital_construction_has_sample_document_comment
                ws4['Q' + str(7 + i)] = object.capital_construction_has_document_template_comment
                ws5['C' + str(7 + i)] = object.preschool_education_id_rgmu
                ws5['D' + str(7 + i)] = object.preschool_education_statement_amount
                ws5['E' + str(7 + i)] = object.preschool_education_link
                ws5['F' + str(7 + i)] = object.preschool_education_has_advanced_appointment_comment
                ws5['G' + str(7 + i)] = object.preschool_education_has_btn_get_service_comment
                ws5['H' + str(7 + i)] = object.preschool_education_has_reglament_comment
                ws5['I' + str(7 + i)] = object.preschool_education_has_estimation_quality_comment
                ws5['J' + str(7 + i)] = object.preschool_education_connected_to_fgis_do_comment
                ws5['K' + str(7 + i)] = object.preschool_education_has_electronic_form_printing_comment
                ws5['L' + str(7 + i)] = object.preschool_education_has_edition_draft_comment
                ws5['M' + str(7 + i)] = object.preschool_education_has_term_of_consideration_comment
                ws5['N' + str(7 + i)] = object.preschool_education_has_notif_consider_result_comment
                ws5['O' + str(7 + i)] = object.preschool_education_has_causes_of_failure_comment
                ws5['P' + str(7 + i)] = object.preschool_education_has_sample_document_comment
                ws5['Q' + str(7 + i)] = object.preschool_education_has_document_template_comment
                ws6['C' + str(7 + i)] = object.school_education_id_rgmu
                ws6['D' + str(7 + i)] = object.school_education_statement_amount
                ws6['E' + str(7 + i)] = object.school_education_link
                ws6['F' + str(7 + i)] = object.school_education_has_advanced_appointment_comment
                ws6['G' + str(7 + i)] = object.school_education_has_btn_get_service_comment
                ws6['H' + str(7 + i)] = object.school_education_has_reglament_comment
                ws6['I' + str(7 + i)] = object.school_education_has_estimation_quality_comment
                ws6['J' + str(7 + i)] = object.school_education_connected_to_fgis_do_comment
                ws6['K' + str(7 + i)] = object.school_education_has_electronic_form_printing_comment
                ws6['L' + str(7 + i)] = object.school_education_has_edition_draft_comment
                ws6['M' + str(7 + i)] = object.school_education_has_term_of_consideration_comment
                ws6['N' + str(7 + i)] = object.school_education_has_notif_consider_result_comment
                ws6['O' + str(7 + i)] = object.school_education_has_causes_of_failure_comment
                ws6['P' + str(7 + i)] = object.school_education_has_sample_document_comment
                ws6['Q' + str(7 + i)] = object.school_education_has_document_template_comment
                ws7['C' + str(7 + i)] = object.needing_premises_id_rgmu
                ws7['D' + str(7 + i)] = object.needing_premises_statement_amount
                ws7['E' + str(7 + i)] = object.needing_premises_link
                ws7['F' + str(7 + i)] = object.needing_premises_has_advanced_appointment_comment
                ws7['G' + str(7 + i)] = object.needing_premises_has_btn_get_service_comment
                ws7['H' + str(7 + i)] = object.needing_premises_has_reglament_comment
                ws7['I' + str(7 + i)] = object.needing_premises_has_estimation_quality_comment
                ws7['J' + str(7 + i)] = object.needing_premises_connected_to_fgis_do_comment
                ws7['K' + str(7 + i)] = object.needing_premises_has_electronic_form_printing_comment
                ws7['L' + str(7 + i)] = object.needing_premises_has_edition_draft_comment
                ws7['M' + str(7 + i)] = object.needing_premises_has_term_of_consideration_comment
                ws7['N' + str(7 + i)] = object.needing_premises_has_notif_consider_result_comment
                ws7['O' + str(7 + i)] = object.needing_premises_has_causes_of_failure_comment
                ws7['P' + str(7 + i)] = object.needing_premises_has_sample_document_comment
                ws7['Q' + str(7 + i)] = object.needing_premises_has_document_template_comment
                ws8['C' + str(7 + i)] = object.town_planning_id_rgmu
                ws8['D' + str(7 + i)] = object.town_planning_statement_amount
                ws8['E' + str(7 + i)] = object.town_planning_link
                ws8['F' + str(7 + i)] = object.town_planning_has_advanced_appointment_comment
                ws8['G' + str(7 + i)] = object.town_planning_has_btn_get_service_comment
                ws8['H' + str(7 + i)] = object.town_planning_has_reglament_comment
                ws8['I' + str(7 + i)] = object.town_planning_has_estimation_quality_comment
                ws8['J' + str(7 + i)] = object.town_planning_connected_to_fgis_do_comment
                ws8['K' + str(7 + i)] = object.town_planning_has_electronic_form_printing_comment
                ws8['L' + str(7 + i)] = object.town_planning_has_edition_draft_comment
                ws8['M' + str(7 + i)] = object.town_planning_has_term_of_consideration_comment
                ws8['N' + str(7 + i)] = object.town_planning_has_notif_consider_result_comment
                ws8['O' + str(7 + i)] = object.town_planning_has_causes_of_failure_comment
                ws8['P' + str(7 + i)] = object.town_planning_has_sample_document_comment
                ws8['Q' + str(7 + i)] = object.town_planning_has_document_template_comment
                ws9['C' + str(7 + i)] = object.archive_reference_id_rgmu
                ws9['D' + str(7 + i)] = object.archive_reference_statement_amount
                ws9['E' + str(7 + i)] = object.archive_reference_link
                ws9['F' + str(7 + i)] = object.archive_reference_has_advanced_appointment_comment
                ws9['G' + str(7 + i)] = object.archive_reference_has_btn_get_service_comment
                ws9['H' + str(7 + i)] = object.archive_reference_has_reglament_comment
                ws9['I' + str(7 + i)] = object.archive_reference_has_estimation_quality_comment
                ws9['J' + str(7 + i)] = object.archive_reference_connected_to_fgis_do_comment
                ws9['K' + str(7 + i)] = object.archive_reference_has_electronic_form_printing_comment
                ws9['L' + str(7 + i)] = object.archive_reference_has_edition_draft_comment
                ws9['M' + str(7 + i)] = object.archive_reference_has_term_of_consideration_comment
                ws9['N' + str(7 + i)] = object.archive_reference_has_notif_consider_result_comment
                ws9['O' + str(7 + i)] = object.archive_reference_has_causes_of_failure_comment
                ws9['P' + str(7 + i)] = object.archive_reference_has_sample_document_comment
                ws9['Q' + str(7 + i)] = object.archive_reference_has_document_template_comment
                ws10['C' + str(7 + i)] = object.land_schemes_id_rgmu
                ws10['D' + str(7 + i)] = object.land_schemes_statement_amount
                ws10['E' + str(7 + i)] = object.land_schemes_link
                ws10['F' + str(7 + i)] = object.land_schemes_has_advanced_appointment_comment
                ws10['G' + str(7 + i)] = object.land_schemes_has_btn_get_service_comment
                ws10['H' + str(7 + i)] = object.land_schemes_has_reglament_comment
                ws10['I' + str(7 + i)] = object.land_schemes_has_estimation_quality_comment
                ws10['J' + str(7 + i)] = object.land_schemes_connected_to_fgis_do_comment
                ws10['K' + str(7 + i)] = object.land_schemes_has_electronic_form_printing_comment
                ws10['L' + str(7 + i)] = object.land_schemes_has_edition_draft_comment
                ws10['M' + str(7 + i)] = object.land_schemes_has_term_of_consideration_comment
                ws10['N' + str(7 + i)] = object.land_schemes_has_notif_consider_result_comment
                ws10['O' + str(7 + i)] = object.land_schemes_has_causes_of_failure_comment
                ws10['P' + str(7 + i)] = object.land_schemes_has_sample_document_comment
                ws10['Q' + str(7 + i)] = object.land_schemes_has_document_template_comment
                ws11['C' + str(7 + i)] = object.land_sale_id_rgmu
                ws11['D' + str(7 + i)] = object.land_sale_statement_amount
                ws11['E' + str(7 + i)] = object.land_sale_link
                ws11['F' + str(7 + i)] = object.land_sale_has_advanced_appointment_comment
                ws11['G' + str(7 + i)] = object.land_sale_has_btn_get_service_comment
                ws11['H' + str(7 + i)] = object.land_sale_has_reglament_comment
                ws11['I' + str(7 + i)] = object.land_sale_has_estimation_quality_comment
                ws11['J' + str(7 + i)] = object.land_sale_connected_to_fgis_do_comment
                ws11['K' + str(7 + i)] = object.land_sale_has_electronic_form_printing_comment
                ws11['L' + str(7 + i)] = object.land_sale_has_edition_draft_comment
                ws11['M' + str(7 + i)] = object.land_sale_has_term_of_consideration_comment
                ws11['N' + str(7 + i)] = object.land_sale_has_notif_consider_result_comment
                ws11['O' + str(7 + i)] = object.land_sale_has_causes_of_failure_comment
                ws11['P' + str(7 + i)] = object.land_sale_has_sample_document_comment
                ws11['Q' + str(7 + i)] = object.land_sale_has_document_template_comment
                ws12['C' + str(7 + i)] = object.land_lease_id_rgmu
                ws12['D' + str(7 + i)] = object.land_lease_statement_amount
                ws12['E' + str(7 + i)] = object.land_lease_link
                ws12['F' + str(7 + i)] = object.land_lease_has_advanced_appointment_comment
                ws12['G' + str(7 + i)] = object.land_lease_has_btn_get_service_comment
                ws12['H' + str(7 + i)] = object.land_lease_has_reglament_comment
                ws12['I' + str(7 + i)] = object.land_lease_has_estimation_quality_comment
                ws12['J' + str(7 + i)] = object.land_lease_connected_to_fgis_do_comment
                ws12['K' + str(7 + i)] = object.land_lease_has_electronic_form_printing_comment
                ws12['L' + str(7 + i)] = object.land_lease_has_edition_draft_comment
                ws12['M' + str(7 + i)] = object.land_lease_has_term_of_consideration_comment
                ws12['N' + str(7 + i)] = object.land_lease_has_notif_consider_result_comment
                ws12['O' + str(7 + i)] = object.land_lease_has_causes_of_failure_comment
                ws12['P' + str(7 + i)] = object.land_lease_has_sample_document_comment
                ws12['Q' + str(7 + i)] = object.land_lease_has_document_template_comment
                ws13['C' + str(7 + i)] = object.ownership_right_id_rgmu
                ws13['D' + str(7 + i)] = object.ownership_right_statement_amount
                ws13['E' + str(7 + i)] = object.ownership_right_link
                ws13['F' + str(7 + i)] = object.ownership_right_has_advanced_appointment_comment
                ws13['G' + str(7 + i)] = object.ownership_right_has_btn_get_service_comment
                ws13['H' + str(7 + i)] = object.ownership_right_has_reglament_comment
                ws13['I' + str(7 + i)] = object.ownership_right_has_estimation_quality_comment
                ws13['J' + str(7 + i)] = object.ownership_right_connected_to_fgis_do_comment
                ws13['K' + str(7 + i)] = object.ownership_right_has_electronic_form_printing_comment
                ws13['L' + str(7 + i)] = object.ownership_right_has_edition_draft_comment
                ws13['M' + str(7 + i)] = object.ownership_right_has_term_of_consideration_comment
                ws13['N' + str(7 + i)] = object.ownership_right_has_notif_consider_result_comment
                ws13['O' + str(7 + i)] = object.ownership_right_has_causes_of_failure_comment
                ws13['P' + str(7 + i)] = object.ownership_right_has_sample_document_comment
                ws13['Q' + str(7 + i)] = object.ownership_right_has_document_template_comment
                ws14['C' + str(7 + i)] = object.municipal_property_lease_id_rgmu
                ws14['D' + str(7 + i)] = object.municipal_property_lease_statement_amount
                ws14['E' + str(7 + i)] = object.municipal_property_lease_link
                ws14['F' + str(7 + i)] = object.municipal_property_lease_has_advanced_appointment_comment
                ws14['G' + str(7 + i)] = object.municipal_property_lease_has_btn_get_service_comment
                ws14['H' + str(7 + i)] = object.municipal_property_lease_has_reglament_comment
                ws14['I' + str(7 + i)] = object.municipal_property_lease_has_estimation_quality_comment
                ws14['J' + str(7 + i)] = object.municipal_property_lease_connected_to_fgis_do_comment
                ws14['K' + str(7 + i)] = object.municipal_property_lease_has_electronic_form_printing_comment
                ws14['L' + str(7 + i)] = object.municipal_property_lease_has_edition_draft_comment
                ws14['M' + str(7 + i)] = object.municipal_property_lease_has_term_of_consideration_comment
                ws14['N' + str(7 + i)] = object.municipal_property_lease_has_notif_consider_result_comment
                ws14['O' + str(7 + i)] = object.municipal_property_lease_has_causes_of_failure_comment
                ws14['P' + str(7 + i)] = object.municipal_property_lease_has_sample_document_comment
                ws14['Q' + str(7 + i)] = object.municipal_property_lease_has_document_template_comment
                ws15['C' + str(7 + i)] = object.free_land_provision_id_rgmu
                ws15['D' + str(7 + i)] = object.free_land_provision_statement_amount
                ws15['E' + str(7 + i)] = object.free_land_provision_link
                ws15['F' + str(7 + i)] = object.free_land_provision_has_advanced_appointment_comment
                ws15['G' + str(7 + i)] = object.free_land_provision_has_btn_get_service_comment
                ws15['H' + str(7 + i)] = object.free_land_provision_has_reglament_comment
                ws15['I' + str(7 + i)] = object.free_land_provision_has_estimation_quality_comment
                ws15['J' + str(7 + i)] = object.free_land_provision_connected_to_fgis_do_comment
                ws15['K' + str(7 + i)] = object.free_land_provision_has_electronic_form_printing_comment
                ws15['L' + str(7 + i)] = object.free_land_provision_has_edition_draft_comment
                ws15['M' + str(7 + i)] = object.free_land_provision_has_term_of_consideration_comment
                ws15['N' + str(7 + i)] = object.free_land_provision_has_notif_consider_result_comment
                ws15['O' + str(7 + i)] = object.free_land_provision_has_causes_of_failure_comment
                ws15['P' + str(7 + i)] = object.free_land_provision_has_sample_document_comment
                ws15['Q' + str(7 + i)] = object.free_land_provision_has_document_template_comment

        response = HttpResponse(save_virtual_workbook(wb),content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}-{}_Examination_Form.xlsx"'.format(year,month)
        return response

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def export_with_troubles(request,year,month):
    if request.method == 'GET':
        wb = load_workbook('./app/static/app/xlsx_templates/with_troubles.xlsx')
        thin_border = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))
        objects = get_with_troubles(month, year)
        ws1 = wb["Лист 1"]
        ws2 = wb["Лист 2"]
        ws3 = wb["Лист 3"]
        ws4 = wb["Лист 4"]
        ws5 = wb["Лист 5"]
        ws6 = wb["Лист 6"]
        ws7 = wb["Лист 7"]
        ws8 = wb["Лист 8"]
        ws9 = wb["Лист 9"]
        ws10 = wb["Лист 10"]
        ws11 = wb["Лист 11"]
        ws12 = wb["Лист 12"]
        ws13 = wb["Лист 13"]
        ws14 = wb["Лист 14"]
        ws15 = wb["Лист 15"]
        ws1[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws1[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[0], month, year)
        ws2[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws2[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[1], month, year)
        ws3[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws3[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[2], month, year)
        ws4[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws4[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[3], month, year)
        ws5[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws5[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[4], month, year)
        ws6[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws6[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[5], month, year)
        ws7[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws7[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[6], month, year)
        ws8[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws8[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[7], month, year)
        ws9[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в (количество указывается нарастающим итогом)'.format(
            month, year)
        ws9[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[8], month, year)
        ws10[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws10[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[9], month, year)
        ws11[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws11[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[10], month, year)
        ws12[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws12[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[11], month, year)
        ws13[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws13[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[12], month, year)
        ws14[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт. (количество указывается нарастающим итогом)'.format(
            month, year)
        ws14[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[13], month, year)
        ws15[
            'D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{}г.,шт.(количество указывается нарастающим итогом)'.format(
            month, year)
        ws15[
            'A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "{}" в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{}г.'.format(
            full_service_names[14], month, year)

        for i in range(len(objects['residential_premises'])):
            object = objects['residential_premises'][i]
            ws1['A1'] = 'Форма самообследования предоставления государственной и муниципальной услуги "Прием заявлений и выдача документов о согласовании проведения переустройства и (или) перепланировки жилого помещения " в электронной форме тех районов Республики Башкортостан, для которых ответ "нет" присутствует в одной из колонок за {}.{} .'.format(month, year)
            ws1['D4'] = 'Количество заявлений поданных на получение государственных/муниципальных услуг в традиционной форме (при личной явке в ведомство) за {}.{} г. ,шт. (количество указывается нарастающим итогом).'
            ws1['A' + str(7 + i)] = i + 1
            ws1['B' + str(7 + i)] = object.region_name
            ws1['C' + str(7 + i)] = object.residential_premises_id_rgmu
            ws1['D' + str(7 + i)] = object.residential_premises_statement_amount
            ws1['E' + str(7 + i)] = object.residential_premises_link
            ws1['F' + str(7 + i)] = object.residential_premises_has_advanced_appointment_comment
            ws1['G' + str(7 + i)] = object.residential_premises_has_btn_get_service_comment
            ws1['H' + str(7 + i)] = object.residential_premises_has_reglament_comment
            ws1['I' + str(7 + i)] = object.residential_premises_has_estimation_quality_comment
            ws1['J' + str(7 + i)] = object.residential_premises_connected_to_fgis_do_comment
            ws1['K' + str(7 + i)] = object.residential_premises_has_electronic_form_printing_comment
            ws1['L' + str(7 + i)] = object.residential_premises_has_edition_draft_comment
            ws1['M' + str(7 + i)] = object.residential_premises_has_term_of_consideration_comment
            ws1['N' + str(7 + i)] = object.residential_premises_has_notif_consider_result_comment
            ws1['O' + str(7 + i)] = object.residential_premises_has_causes_of_failure_comment
            ws1['P' + str(7 + i)] = object.residential_premises_has_sample_document_comment
            ws1['Q' + str(7 + i)] = object.residential_premises_has_document_template_comment
        for i in range(len(objects['housing_transfer'])):
            object = objects['housing_transfer'][i]
            ws2['A' + str(7+i)] = i+1
            ws2['B' + str(7+i)] = object.region_name
            ws2['C' + str(7 + i)] = object.housing_transfer_id_rgmu
            ws2['D' + str(7 + i)] = object.housing_transfer_statement_amount
            ws2['E' + str(7 + i)] = object.housing_transfer_link
            ws2['F' + str(7 + i)] = object.housing_transfer_has_advanced_appointment_comment
            ws2['G' + str(7 + i)] = object.housing_transfer_has_btn_get_service_comment
            ws2['H' + str(7 + i)] = object.housing_transfer_has_reglament_comment
            ws2['I' + str(7 + i)] = object.housing_transfer_has_estimation_quality_comment
            ws2['J' + str(7 + i)] = object.housing_transfer_connected_to_fgis_do_comment
            ws2['K' + str(7 + i)] = object.housing_transfer_has_electronic_form_printing_comment
            ws2['L' + str(7 + i)] = object.housing_transfer_has_edition_draft_comment
            ws2['M' + str(7 + i)] = object.housing_transfer_has_term_of_consideration_comment
            ws2['N' + str(7 + i)] = object.housing_transfer_has_notif_consider_result_comment
            ws2['O' + str(7 + i)] = object.housing_transfer_has_causes_of_failure_comment
            ws2['P' + str(7 + i)] = object.housing_transfer_has_sample_document_comment
            ws2['Q' + str(7 + i)] = object.housing_transfer_has_document_template_comment
        for i in range(len(objects['advertising_structures'])):
            object = objects['advertising_structures'][i]
            ws3['A' + str(7 + i)] = i + 1
            ws3['B' + str(7 + i)] = object.region_name
            ws3['C' + str(7 + i)] = object.advertising_structures_id_rgmu
            ws3['D' + str(7 + i)] = object.advertising_structures_statement_amount
            ws3['E' + str(7 + i)] = object.advertising_structures_link
            ws3['F' + str(7 + i)] = object.advertising_structures_has_advanced_appointment_comment
            ws3['G' + str(7 + i)] = object.advertising_structures_has_btn_get_service_comment
            ws3['H' + str(7 + i)] = object.advertising_structures_has_reglament_comment
            ws3['I' + str(7 + i)] = object.advertising_structures_has_estimation_quality_comment
            ws3['J' + str(7 + i)] = object.advertising_structures_connected_to_fgis_do_comment
            ws3['K' + str(7 + i)] = object.advertising_structures_has_electronic_form_printing_comment
            ws3['L' + str(7 + i)] = object.advertising_structures_has_edition_draft_comment
            ws3['M' + str(7 + i)] = object.advertising_structures_has_term_of_consideration_comment
            ws3['N' + str(7 + i)] = object.advertising_structures_has_notif_consider_result_comment
            ws3['O' + str(7 + i)] = object.advertising_structures_has_causes_of_failure_comment
            ws3['P' + str(7 + i)] = object.advertising_structures_has_sample_document_comment
            ws3['Q' + str(7 + i)] = object.advertising_structures_has_document_template_comment
        for i in range(len(objects['capital_construction'])):
            object = objects['capital_construction'][i]
            ws4['A' + str(7 + i)] = i + 1
            ws4['B' + str(7 + i)] = object.region_name
            ws4['C' + str(7 + i)] = object.capital_construction_id_rgmu
            ws4['D' + str(7 + i)] = object.capital_construction_statement_amount
            ws4['E' + str(7 + i)] = object.capital_construction_link
            ws4['F' + str(7 + i)] = object.capital_construction_has_advanced_appointment_comment
            ws4['G' + str(7 + i)] = object.capital_construction_has_btn_get_service_comment
            ws4['H' + str(7 + i)] = object.capital_construction_has_reglament_comment
            ws4['I' + str(7 + i)] = object.capital_construction_has_estimation_quality_comment
            ws4['J' + str(7 + i)] = object.capital_construction_connected_to_fgis_do_comment
            ws4['K' + str(7 + i)] = object.capital_construction_has_electronic_form_printing_comment
            ws4['L' + str(7 + i)] = object.capital_construction_has_edition_draft_comment
            ws4['M' + str(7 + i)] = object.capital_construction_has_term_of_consideration_comment
            ws4['N' + str(7 + i)] = object.capital_construction_has_notif_consider_result_comment
            ws4['O' + str(7 + i)] = object.capital_construction_has_causes_of_failure_comment
            ws4['P' + str(7 + i)] = object.capital_construction_has_sample_document_comment
            ws4['Q' + str(7 + i)] = object.capital_construction_has_document_template_comment
        for i in range(len(objects['preschool_education'])):
            object = objects['preschool_education'][i]
            ws5['A' + str(7 + i)] = i + 1
            ws5['B' + str(7 + i)] = object.region_name
            ws5['C' + str(7 + i)] = object.preschool_education_id_rgmu
            ws5['D' + str(7 + i)] = object.preschool_education_statement_amount
            ws5['E' + str(7 + i)] = object.preschool_education_link
            ws5['F' + str(7 + i)] = object.preschool_education_has_advanced_appointment_comment
            ws5['G' + str(7 + i)] = object.preschool_education_has_btn_get_service_comment
            ws5['H' + str(7 + i)] = object.preschool_education_has_reglament_comment
            ws5['I' + str(7 + i)] = object.preschool_education_has_estimation_quality_comment
            ws5['J' + str(7 + i)] = object.preschool_education_connected_to_fgis_do_comment
            ws5['K' + str(7 + i)] = object.preschool_education_has_electronic_form_printing_comment
            ws5['L' + str(7 + i)] = object.preschool_education_has_edition_draft_comment
            ws5['M' + str(7 + i)] = object.preschool_education_has_term_of_consideration_comment
            ws5['N' + str(7 + i)] = object.preschool_education_has_notif_consider_result_comment
            ws5['O' + str(7 + i)] = object.preschool_education_has_causes_of_failure_comment
            ws5['P' + str(7 + i)] = object.preschool_education_has_sample_document_comment
            ws5['Q' + str(7 + i)] = object.preschool_education_has_document_template_comment
        for i in range(len(objects['school_education'])):
            object = objects['school_education'][i]
            ws6['A' + str(7 + i)] = i + 1
            ws6['B' + str(7 + i)] = object.region_name
            ws6['C' + str(7 + i)] = object.school_education_id_rgmu
            ws6['D' + str(7 + i)] = object.school_education_statement_amount
            ws6['E' + str(7 + i)] = object.school_education_link
            ws6['F' + str(7 + i)] = object.school_education_has_advanced_appointment_comment
            ws6['G' + str(7 + i)] = object.school_education_has_btn_get_service_comment
            ws6['H' + str(7 + i)] = object.school_education_has_reglament_comment
            ws6['I' + str(7 + i)] = object.school_education_has_estimation_quality_comment
            ws6['J' + str(7 + i)] = object.school_education_connected_to_fgis_do_comment
            ws6['K' + str(7 + i)] = object.school_education_has_electronic_form_printing_comment
            ws6['L' + str(7 + i)] = object.school_education_has_edition_draft_comment
            ws6['M' + str(7 + i)] = object.school_education_has_term_of_consideration_comment
            ws6['N' + str(7 + i)] = object.school_education_has_notif_consider_result_comment
            ws6['O' + str(7 + i)] = object.school_education_has_causes_of_failure_comment
            ws6['P' + str(7 + i)] = object.school_education_has_sample_document_comment
            ws6['Q' + str(7 + i)] = object.school_education_has_document_template_comment
        for i in range(len(objects['needing_premises'])):
            object = objects['needing_premises'][i]
            ws7['A' + str(7 + i)] = i + 1
            ws7['B' + str(7 + i)] = object.region_name
            ws7['C' + str(7 + i)] = object.needing_premises_id_rgmu
            ws7['D' + str(7 + i)] = object.needing_premises_statement_amount
            ws7['E' + str(7 + i)] = object.needing_premises_link
            ws7['F' + str(7 + i)] = object.needing_premises_has_advanced_appointment_comment
            ws7['G' + str(7 + i)] = object.needing_premises_has_btn_get_service_comment
            ws7['H' + str(7 + i)] = object.needing_premises_has_reglament_comment
            ws7['I' + str(7 + i)] = object.needing_premises_has_estimation_quality_comment
            ws7['J' + str(7 + i)] = object.needing_premises_connected_to_fgis_do_comment
            ws7['K' + str(7 + i)] = object.needing_premises_has_electronic_form_printing_comment
            ws7['L' + str(7 + i)] = object.needing_premises_has_edition_draft_comment
            ws7['M' + str(7 + i)] = object.needing_premises_has_term_of_consideration_comment
            ws7['N' + str(7 + i)] = object.needing_premises_has_notif_consider_result_comment
            ws7['O' + str(7 + i)] = object.needing_premises_has_causes_of_failure_comment
            ws7['P' + str(7 + i)] = object.needing_premises_has_sample_document_comment
            ws7['Q' + str(7 + i)] = object.needing_premises_has_document_template_comment
        for i in range(len(objects['town_planning'])):
            object = objects['town_planning'][i]
            ws8['A' + str(7 + i)] = i + 1
            ws8['B' + str(7 + i)] = object.region_name
            ws8['C' + str(7 + i)] = object.town_planning_id_rgmu
            ws8['D' + str(7 + i)] = object.town_planning_statement_amount
            ws8['E' + str(7 + i)] = object.town_planning_link
            ws8['F' + str(7 + i)] = object.town_planning_has_advanced_appointment_comment
            ws8['G' + str(7 + i)] = object.town_planning_has_btn_get_service_comment
            ws8['H' + str(7 + i)] = object.town_planning_has_reglament_comment
            ws8['I' + str(7 + i)] = object.town_planning_has_estimation_quality_comment
            ws8['J' + str(7 + i)] = object.town_planning_connected_to_fgis_do_comment
            ws8['K' + str(7 + i)] = object.town_planning_has_electronic_form_printing_comment
            ws8['L' + str(7 + i)] = object.town_planning_has_edition_draft_comment
            ws8['M' + str(7 + i)] = object.town_planning_has_term_of_consideration_comment
            ws8['N' + str(7 + i)] = object.town_planning_has_notif_consider_result_comment
            ws8['O' + str(7 + i)] = object.town_planning_has_causes_of_failure_comment
            ws8['P' + str(7 + i)] = object.town_planning_has_sample_document_comment
            ws8['Q' + str(7 + i)] = object.town_planning_has_document_template_comment
        for i in range(len(objects['archive_reference'])):
            object = objects['archive_reference'][i]
            ws9['A' + str(7 + i)] = i + 1
            ws9['B' + str(7 + i)] = object.region_name
            ws9['C' + str(7 + i)] = object.archive_reference_id_rgmu
            ws9['D' + str(7 + i)] = object.archive_reference_statement_amount
            ws9['E' + str(7 + i)] = object.archive_reference_link
            ws9['F' + str(7 + i)] = object.archive_reference_has_advanced_appointment_comment
            ws9['G' + str(7 + i)] = object.archive_reference_has_btn_get_service_comment
            ws9['H' + str(7 + i)] = object.archive_reference_has_reglament_comment
            ws9['I' + str(7 + i)] = object.archive_reference_has_estimation_quality_comment
            ws9['J' + str(7 + i)] = object.archive_reference_connected_to_fgis_do_comment
            ws9['K' + str(7 + i)] = object.archive_reference_has_electronic_form_printing_comment
            ws9['L' + str(7 + i)] = object.archive_reference_has_edition_draft_comment
            ws9['M' + str(7 + i)] = object.archive_reference_has_term_of_consideration_comment
            ws9['N' + str(7 + i)] = object.archive_reference_has_notif_consider_result_comment
            ws9['O' + str(7 + i)] = object.archive_reference_has_causes_of_failure_comment
            ws9['P' + str(7 + i)] = object.archive_reference_has_sample_document_comment
            ws9['Q' + str(7 + i)] = object.archive_reference_has_document_template_comment
        for i in range(len(objects['land_schemes'])):
            object = objects['land_schemes'][i]
            ws10['A' + str(7 + i)] = i + 1
            ws10['B' + str(7 + i)] = object.region_name
            ws10['C' + str(7 + i)] = object.land_schemes_id_rgmu
            ws10['D' + str(7 + i)] = object.land_schemes_statement_amount
            ws10['E' + str(7 + i)] = object.land_schemes_link
            ws10['F' + str(7 + i)] = object.land_schemes_has_advanced_appointment_comment
            ws10['G' + str(7 + i)] = object.land_schemes_has_btn_get_service_comment
            ws10['H' + str(7 + i)] = object.land_schemes_has_reglament_comment
            ws10['I' + str(7 + i)] = object.land_schemes_has_estimation_quality_comment
            ws10['J' + str(7 + i)] = object.land_schemes_connected_to_fgis_do_comment
            ws10['K' + str(7 + i)] = object.land_schemes_has_electronic_form_printing_comment
            ws10['L' + str(7 + i)] = object.land_schemes_has_edition_draft_comment
            ws10['M' + str(7 + i)] = object.land_schemes_has_term_of_consideration_comment
            ws10['N' + str(7 + i)] = object.land_schemes_has_notif_consider_result_comment
            ws10['O' + str(7 + i)] = object.land_schemes_has_causes_of_failure_comment
            ws10['P' + str(7 + i)] = object.land_schemes_has_sample_document_comment
            ws10['Q' + str(7 + i)] = object.land_schemes_has_document_template_comment
        for i in range(len(objects['land_sale'])):
            object = objects['land_sale'][i]
            ws11['A' + str(7 + i)] = i + 1
            ws11['B' + str(7 + i)] = object.region_name
            ws11['C' + str(7 + i)] = object.land_sale_id_rgmu
            ws11['D' + str(7 + i)] = object.land_sale_statement_amount
            ws11['E' + str(7 + i)] = object.land_sale_link
            ws11['F' + str(7 + i)] = object.land_sale_has_advanced_appointment_comment
            ws11['G' + str(7 + i)] = object.land_sale_has_btn_get_service_comment
            ws11['H' + str(7 + i)] = object.land_sale_has_reglament_comment
            ws11['I' + str(7 + i)] = object.land_sale_has_estimation_quality_comment
            ws11['J' + str(7 + i)] = object.land_sale_connected_to_fgis_do_comment
            ws11['K' + str(7 + i)] = object.land_sale_has_electronic_form_printing_comment
            ws11['L' + str(7 + i)] = object.land_sale_has_edition_draft_comment
            ws11['M' + str(7 + i)] = object.land_sale_has_term_of_consideration_comment
            ws11['N' + str(7 + i)] = object.land_sale_has_notif_consider_result_comment
            ws11['O' + str(7 + i)] = object.land_sale_has_causes_of_failure_comment
            ws11['P' + str(7 + i)] = object.land_sale_has_sample_document_comment
            ws11['Q' + str(7 + i)] = object.land_sale_has_document_template_comment
        for i in range(len(objects['land_lease'])):
            object = objects['land_lease'][i]
            ws12['A' + str(7+i)] = i+1
            ws12['B' + str(7+i)] = object.region_name
            ws12['C' + str(7 + i)] = object.land_lease_id_rgmu
            ws12['D' + str(7 + i)] = object.land_lease_statement_amount
            ws12['E' + str(7 + i)] = object.land_lease_link
            ws12['F' + str(7 + i)] = object.land_lease_has_advanced_appointment_comment
            ws12['G' + str(7 + i)] = object.land_lease_has_btn_get_service_comment
            ws12['H' + str(7 + i)] = object.land_lease_has_reglament_comment
            ws12['I' + str(7 + i)] = object.land_lease_has_estimation_quality_comment
            ws12['J' + str(7 + i)] = object.land_lease_connected_to_fgis_do_comment
            ws12['K' + str(7 + i)] = object.land_lease_has_electronic_form_printing_comment
            ws12['L' + str(7 + i)] = object.land_lease_has_edition_draft_comment
            ws12['M' + str(7 + i)] = object.land_lease_has_term_of_consideration_comment
            ws12['N' + str(7 + i)] = object.land_lease_has_notif_consider_result_comment
            ws12['O' + str(7 + i)] = object.land_lease_has_causes_of_failure_comment
            ws12['P' + str(7 + i)] = object.land_lease_has_sample_document_comment
            ws12['Q' + str(7 + i)] = object.land_lease_has_document_template_comment
        for i in range(len(objects['ownership_right'])):
            object = objects['ownership_right'][i]
            ws13['A' + str(7 + i)] = i + 1
            ws13['B' + str(7 + i)] = object.region_name
            ws13['C' + str(7 + i)] = object.ownership_right_id_rgmu
            ws13['D' + str(7 + i)] = object.ownership_right_statement_amount
            ws13['E' + str(7 + i)] = object.ownership_right_link
            ws13['F' + str(7 + i)] = object.ownership_right_has_advanced_appointment_comment
            ws13['G' + str(7 + i)] = object.ownership_right_has_btn_get_service_comment
            ws13['H' + str(7 + i)] = object.ownership_right_has_reglament_comment
            ws13['I' + str(7 + i)] = object.ownership_right_has_estimation_quality_comment
            ws13['J' + str(7 + i)] = object.ownership_right_connected_to_fgis_do_comment
            ws13['K' + str(7 + i)] = object.ownership_right_has_electronic_form_printing_comment
            ws13['L' + str(7 + i)] = object.ownership_right_has_edition_draft_comment
            ws13['M' + str(7 + i)] = object.ownership_right_has_term_of_consideration_comment
            ws13['N' + str(7 + i)] = object.ownership_right_has_notif_consider_result_comment
            ws13['O' + str(7 + i)] = object.ownership_right_has_causes_of_failure_comment
            ws13['P' + str(7 + i)] = object.ownership_right_has_sample_document_comment
            ws13['Q' + str(7 + i)] = object.ownership_right_has_document_template_comment
        for i in range(len(objects['municipal_property_lease'])):
            object = objects['municipal_property_lease'][i]
            ws14['A' + str(7 + i)] = i + 1
            ws14['B' + str(7 + i)] = object.region_name
            ws14['C' + str(7 + i)] = object.municipal_property_lease_id_rgmu
            ws14['D' + str(7 + i)] = object.municipal_property_lease_statement_amount
            ws14['E' + str(7 + i)] = object.municipal_property_lease_link
            ws14['F' + str(7 + i)] = object.municipal_property_lease_has_advanced_appointment_comment
            ws14['G' + str(7 + i)] = object.municipal_property_lease_has_btn_get_service_comment
            ws14['H' + str(7 + i)] = object.municipal_property_lease_has_reglament_comment
            ws14['I' + str(7 + i)] = object.municipal_property_lease_has_estimation_quality_comment
            ws14['J' + str(7 + i)] = object.municipal_property_lease_connected_to_fgis_do_comment
            ws14['K' + str(7 + i)] = object.municipal_property_lease_has_electronic_form_printing_comment
            ws14['L' + str(7 + i)] = object.municipal_property_lease_has_edition_draft_comment
            ws14['M' + str(7 + i)] = object.municipal_property_lease_has_term_of_consideration_comment
            ws14['N' + str(7 + i)] = object.municipal_property_lease_has_notif_consider_result_comment
            ws14['O' + str(7 + i)] = object.municipal_property_lease_has_causes_of_failure_comment
            ws14['P' + str(7 + i)] = object.municipal_property_lease_has_sample_document_comment
            ws14['Q' + str(7 + i)] = object.municipal_property_lease_has_document_template_comment
        for i in range(len(objects['free_land_provision'])):
            object = objects['free_land_provision'][i]
            ws15['A' + str(7 + i)] = i + 1
            ws15['B' + str(7 + i)] = object.region_name
            ws15['C' + str(7 + i)] = object.free_land_provision_id_rgmu
            ws15['D' + str(7 + i)] = object.free_land_provision_statement_amount
            ws15['E' + str(7 + i)] = object.free_land_provision_link
            ws15['F' + str(7 + i)] = object.free_land_provision_has_advanced_appointment_comment
            ws15['G' + str(7 + i)] = object.free_land_provision_has_btn_get_service_comment
            ws15['H' + str(7 + i)] = object.free_land_provision_has_reglament_comment
            ws15['I' + str(7 + i)] = object.free_land_provision_has_estimation_quality_comment
            ws15['J' + str(7 + i)] = object.free_land_provision_connected_to_fgis_do_comment
            ws15['K' + str(7 + i)] = object.free_land_provision_has_electronic_form_printing_comment
            ws15['L' + str(7 + i)] = object.free_land_provision_has_edition_draft_comment
            ws15['M' + str(7 + i)] = object.free_land_provision_has_term_of_consideration_comment
            ws15['N' + str(7 + i)] = object.free_land_provision_has_notif_consider_result_comment
            ws15['O' + str(7 + i)] = object.free_land_provision_has_causes_of_failure_comment
            ws15['P' + str(7 + i)] = object.free_land_provision_has_sample_document_comment
            ws15['Q' + str(7 + i)] = object.free_land_provision_has_document_template_comment
        for _row in ws1.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws1.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws2.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws2.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws3.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws3.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws4.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws4.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws5.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws5.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws6.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws6.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws7.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws7.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws8.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws8.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws9.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws9.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws10.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws10.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws11.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws11.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws12.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws12.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws13.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws13.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws14.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws14.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        for _row in ws15.iter_rows(min_row=7,min_col=1,max_col=17,max_row=ws15.max_row):
            for _cell in _row:
                _cell.border = thin_border
                _cell.alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}-{}_With_Troubles_Form.xlsx"'.format(year, month)
        return response

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def export_not_sent(request,year,month):
    if request.method == 'GET':
        wb = load_workbook('./app/static/app/xlsx_templates/not_sent.xlsx')
        objects = get_not_sent(month, year)
        ws = wb["Лист1"]
        ws['A1'] = 'Список районов , не заполнивших форму за период {}.{} г.'.format(month,year)
        for i in range(len(objects)):
            ws['A'+str(3+i)] = objects[i]
        response = HttpResponse(save_virtual_workbook(wb),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="{}-{}_Not_Sent_Form.xlsx"'.format(year, month)
        return response

@login_required(login_url='/login/',
                redirect_field_name='/result_form/residential_premises/' + datetime.today().strftime('%Y/%m/'))
def empty_view(request):
    return HttpResponseRedirect('/login/')

@login_required(login_url='/login/',redirect_field_name='/profile/')
def get_profile(request):
    if request.method == 'GET':
        return render(request, 'app/profile.html', {
                                           'username': request.user.username,
                                            'region_name': request.user.region_name,
                                            'email': request.user.email,
                                           'year': datetime.today().strftime('%Y'),
                                           'month': MONTHS[MONTH_NUMBERS.index(datetime.today().strftime('%m'))],
                                           'num_month': datetime.today().strftime('%m'),
                                           'zipped': zip(regions_names, short_regions_names),
                                           'years': [i for i in range(2016, int(datetime.now().year) + 1)],
                                           'zipped_service_names': zip(full_service_names, short_service_names)
                                           })
    else:
        return HttpResponseNotFound('Sorry Page Not Found')



@login_required(login_url='/login/',redirect_field_name='/profile/edit/')
def edit_profile(request):
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST,instance=request.user)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect('/profile/')
    elif request.method == 'GET':
        form = CustomUserChangeForm(instance=request.user)

    return render(request, 'app/edit_profile.html', {
                                            'region_name': request.user.region_name,
                                           'form': form,
                                           'year': datetime.today().strftime('%Y'),
                                           'month': MONTHS[MONTH_NUMBERS.index(datetime.today().strftime('%m'))],
                                           'num_month': datetime.today().strftime('%m'),
                                           'zipped': zip(regions_names, short_regions_names),
                                           'years': [i for i in range(2016, int(datetime.now().year) + 1)],
                                           'zipped_service_names': zip(full_service_names, short_service_names)
                                           })